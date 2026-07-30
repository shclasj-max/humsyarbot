"""
📥 excel_report — ساخت فایل اکسل کامل دیتابیس.

این منطق قبلاً فقط داخل admin.py (دکمه پنل ربات) بود و سمت web
هیچ‌وقت مصرف نمی‌شد. حالا به‌صورت ماژول مشترک است تا هر دو کانال
(پنل ربات + درخواست از مینی‌اپ از طریق صف bot_notifications)
به یک خروجی یکسان برسند.
"""
import io
import logging
from datetime import datetime

from database import db
from utils import fmt_jalali_dt

logger = logging.getLogger(__name__)


async def build_database_excel():
    """خروجی کامل دیتابیس (کاربران، تیکت‌ها، بانک سؤال) با سه شیت.

    برمی‌گرداند: (buffer, filename, caption, counts_dict)
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill

    wb = openpyxl.Workbook()
    header_fill = PatternFill(
        start_color='1F4E78', end_color='1F4E78', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True)

    # ── شیت کاربران ──
    ws = wb.active
    ws.title = 'کاربران'
    headers = ['آیدی', 'نام', 'شماره دانشجویی', 'گروه', 'ورودی', 'یوزرنیم',
               'وضعیت', 'تاریخ ثبت‌نام', 'تعداد پاسخ', 'پاسخ صحیح']
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    users = await db.all_users(approved_only=False)
    for u in users:
        ws.append([
            u.get('user_id', ''), u.get('name', ''),
            u.get('student_id', '') or '—',
            u.get('group', ''), u.get('intake', '') or '—',
            u.get('username', '') or '—',
            'تأییدشده' if u.get('approved') else 'در انتظار',
            fmt_jalali_dt(u.get('registered_at', ''), with_time=False),
            u.get('total_answers', 0), u.get('correct_answers', 0),
        ])

    # ── شیت تیکت‌ها ──
    ws2 = wb.create_sheet('تیکت‌ها')
    headers2 = ['شماره تیکت', 'کاربر', 'موضوع', 'وضعیت', 'تاریخ ثبت']
    ws2.append(headers2)
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
    tickets = await db.tickets.find({}).sort('created_at', -1).to_list(2000)
    for t in tickets:
        ws2.append([
            t.get('ticket_id', ''), t.get('user_name', ''), t.get('subject', ''),
            'باز' if t.get('status') == 'open' else 'بسته',
            fmt_jalali_dt(t.get('created_at', ''), with_time=False),
        ])

    # ── شیت بانک سؤال ──
    ws3 = wb.create_sheet('بانک سوال')
    headers3 = ['درس', 'مبحث', 'سختی', 'تعداد پاسخ', 'پاسخ صحیح', 'وضعیت تأیید']
    ws3.append(headers3)
    for cell in ws3[1]:
        cell.fill = header_fill
        cell.font = header_font
    questions = await db.questions.find({}).to_list(5000)
    for q in questions:
        ws3.append([
            q.get('lesson', ''), q.get('topic', ''), q.get('difficulty', ''),
            q.get('attempt_count', 0), q.get('correct_count', 0),
            'تأییدشده' if q.get('approved') else 'در انتظار',
        ])

    # عرض خودکار ستون‌ها
    for sheet in wb.worksheets:
        for col in sheet.columns:
            max_len = max(
                (len(str(c.value)) for c in col if c.value is not None),
                default=10)
            sheet.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    fname = f"humsyar_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    caption = (f"📥 <b>خروجی کامل دیتابیس</b>\n"
               f"👥 {len(users)} کاربر | 🎫 {len(tickets)} تیکت | "
               f"🧪 {len(questions)} سوال")
    counts = {
        "users": len(users),
        "tickets": len(tickets),
        "questions": len(questions),
    }
    return buf, fname, caption, counts
