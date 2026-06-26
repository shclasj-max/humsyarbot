"""
👨‍⚕️ پنل ادمین — نسخه کامل و حرفه‌ای
  ✅ broadcast پیشرفته: preview + تأیید + ارسال به گروه خاص + ارسال زماندار
  ✅ فیکس باگ duplicate key در restore بکاپ
  ✅ فیکس سرچ کاربران
  ✅ pagination و filter کاربران
"""
import os
import asyncio
import logging
from datetime import datetime, timedelta
from telegram import (
    Update, InlineKeyboardButton, InlineKeyboardMarkup,
    Message
)
from telegram.ext import ContextTypes, ConversationHandler
from database import db
from utils import main_keyboard, content_admin_keyboard, admin_keyboard, safe_send, send_audit_log

logger   = logging.getLogger(__name__)
ADMIN_ID = int(os.getenv('ADMIN_ID', '0'))
BROADCAST = 5   # نگه داشته شده برای سازگاری با bot.py


async def _admin_menu(query_or_msg, edit: bool = True, uid: int = None):
    """
    FIX جدید: منو بر اساس نقش فیلتر می‌شود —
    ADMIN_ID همه‌چیز می‌بیند، نقش‌های فرعی فقط بخش مجاز خودشان.
    """
    s     = await db.global_stats()
    role  = None
    if uid is not None and uid != ADMIN_ID:
        role_doc = await db.get_admin_role(uid)
        role = role_doc.get('role') if role_doc else None

    # نقش پشتیبان: منوی بسیار محدود
    if role == 'support':
        keyboard = [
            [InlineKeyboardButton("🎫 مدیریت تیکت‌ها", callback_data='ticket:admin_list')],
        ]
        text = "🎫 <b>پنل پشتیبان</b>\n━━━━━━━━━━━━━━━━\nشما فقط به مدیریت تیکت‌ها دسترسی دارید."
        markup = InlineKeyboardMarkup(keyboard)
        try:
            if edit and hasattr(query_or_msg, 'edit_message_text'):
                await query_or_msg.edit_message_text(text, parse_mode='HTML', reply_markup=markup)
            else:
                msg = query_or_msg if hasattr(query_or_msg, 'reply_text') else query_or_msg.message
                await msg.reply_text(text, parse_mode='HTML', reply_markup=markup)
        except Exception as e:
            logger.debug(f"_admin_menu(support): {e}")
        return

    # نقش مسئول اطلاعیه: فقط broadcast
    if role == 'broadcaster':
        keyboard = [
            [InlineKeyboardButton("📢 ارسال همگانی", callback_data='admin:broadcast')],
        ]
        text = "📢 <b>پنل مسئول اطلاعیه</b>\n━━━━━━━━━━━━━━━━"
        markup = InlineKeyboardMarkup(keyboard)
        try:
            if edit and hasattr(query_or_msg, 'edit_message_text'):
                await query_or_msg.edit_message_text(text, parse_mode='HTML', reply_markup=markup)
            else:
                msg = query_or_msg if hasattr(query_or_msg, 'reply_text') else query_or_msg.message
                await msg.reply_text(text, parse_mode='HTML', reply_markup=markup)
        except Exception as e:
            logger.debug(f"_admin_menu(broadcaster): {e}")
        return

    # FIX امنیتی: content_scoped باید به پنل محتوا هدایت شود،
    # نه منوی کامل ادمین ارشد را ببیند.
    if role == 'content_scoped':
        keyboard = [
            [InlineKeyboardButton("🎓 رفتن به پنل محتوا", callback_data='ca:main')],
        ]
        text = (
            "📅 <b>پنل مدیر محتوای محدود</b>\n━━━━━━━━━━━━━━━━\n\n"
            "شما فقط به محتوای ورودی خاص خودتان دسترسی دارید.\n"
            "از دکمه زیر وارد پنل محتوا شوید:"
        )
        markup = InlineKeyboardMarkup(keyboard)
        try:
            if edit and hasattr(query_or_msg, 'edit_message_text'):
                await query_or_msg.edit_message_text(text, parse_mode='HTML', reply_markup=markup)
            else:
                msg = query_or_msg if hasattr(query_or_msg, 'reply_text') else query_or_msg.message
                await msg.reply_text(text, parse_mode='HTML', reply_markup=markup)
        except Exception as e:
            logger.debug(f"_admin_menu(content_scoped): {e}")
        return

    # FIX جدید: نقش خرخون — فقط دسترسی به بررسی گزارش سوال/جزوه
    if role == 'reviewer':
        keyboard = [
            [InlineKeyboardButton("⚠️ گزارشات سوال/جزوه", callback_data='report:manage:all')],
        ]
        text = "🤓 <b>پنل خرخون</b>\n━━━━━━━━━━━━━━━━\nشما به بررسی گزارشات سوال و جزوه دسترسی دارید."
        markup = InlineKeyboardMarkup(keyboard)
        try:
            if edit and hasattr(query_or_msg, 'edit_message_text'):
                await query_or_msg.edit_message_text(text, parse_mode='HTML', reply_markup=markup)
            else:
                msg = query_or_msg if hasattr(query_or_msg, 'reply_text') else query_or_msg.message
                await msg.reply_text(text, parse_mode='HTML', reply_markup=markup)
        except Exception as e:
            logger.debug(f"_admin_menu(reviewer): {e}")
        return

    # FIX جدید: ادمین ربات (نماینده) — دسترسی به کاربران، برنامه‌ها، اطلاعیه‌ها
    if role == 'bot_admin':
        keyboard = [
            [InlineKeyboardButton("👥 مدیریت کاربران",  callback_data='admin:users:0')],
            [
                InlineKeyboardButton("📅 برنامه جدید",  callback_data='schedule:add_type'),
                InlineKeyboardButton("🗑 حذف برنامه",   callback_data='schedule:del_list'),
            ],
            [InlineKeyboardButton("📢 ارسال همگانی",    callback_data='admin:broadcast')],
        ]
        text = "👮 <b>پنل ادمین ربات (نماینده)</b>\n━━━━━━━━━━━━━━━━\nدسترسی محدود — بدون تنظیمات حیاتی."
        markup = InlineKeyboardMarkup(keyboard)
        try:
            if edit and hasattr(query_or_msg, 'edit_message_text'):
                await query_or_msg.edit_message_text(text, parse_mode='HTML', reply_markup=markup)
            else:
                msg = query_or_msg if hasattr(query_or_msg, 'reply_text') else query_or_msg.message
                await msg.reply_text(text, parse_mode='HTML', reply_markup=markup)
        except Exception as e:
            logger.debug(f"_admin_menu(bot_admin): {e}")
        return

    # فقط ادمین ارشد به این نقطه می‌رسد — منوی کامل
    keyboard = [
        [InlineKeyboardButton(
            f"📊 آمار سیستم  ({s['users']} کاربر | {s.get('open_tickets', 0)} تیکت باز)",
            callback_data='admin:stats'
        )],
        [
            InlineKeyboardButton("👥 مدیریت کاربران",  callback_data='admin:users:0'),
            InlineKeyboardButton("⏳ تأیید کاربران",   callback_data='admin:pending'),
        ],
        [InlineKeyboardButton("📅 مدیریت ورودی‌ها",   callback_data='admin:intakes')],
        [InlineKeyboardButton("🔍 جستجوی کاربر",       callback_data='admin:search_user')],
        [InlineKeyboardButton("🎓 ادمین‌های محتوا",    callback_data='admin:content_admins')],
        [
            InlineKeyboardButton("📘 علوم پایه",       callback_data='ca:terms_admin'),
            InlineKeyboardButton("📚 رفرنس‌ها",         callback_data='ca:refs_admin'),
        ],
        [InlineKeyboardButton("❓ مدیریت FAQ",          callback_data='ca:faq')],
        [
            InlineKeyboardButton("🧪 بانک سوال",       callback_data='admin:qbank_manage'),
            InlineKeyboardButton("✅ تأیید سوالات",    callback_data='admin:pending_q'),
        ],
        [
            InlineKeyboardButton("📅 برنامه جدید",     callback_data='schedule:add_type'),
            InlineKeyboardButton("🗑 حذف برنامه",      callback_data='schedule:del_list'),
        ],
        [InlineKeyboardButton("🔄 اعلام تغییر زمان (کلاس منعطف)", callback_data='schedule:flex_list')],
        [InlineKeyboardButton("🎫 مدیریت تیکت‌ها",     callback_data='ticket:admin_list')],
        [InlineKeyboardButton("⚠️ گزارشات سوال/جزوه",  callback_data='report:manage:all')],
        [InlineKeyboardButton("📢 ارسال همگانی",        callback_data='admin:broadcast')],
        [InlineKeyboardButton("💾 پشتیبان‌گیری",        callback_data='backup:menu')],
        [InlineKeyboardButton("📡 وضعیت ربات",            callback_data='admin:bot_status')],
    ]
    # FIX جدید: دکمه‌های ویژه ادمین ارشد — نقش‌های فرعی این‌ها را نمی‌بینند
    if uid is None or uid == ADMIN_ID:
        keyboard.append([
            InlineKeyboardButton("🛡 سطوح دسترسی ادمین", callback_data='admin:roles'),
            InlineKeyboardButton("⚙️ تنظیمات ربات",      callback_data='admin:settings'),
        ])
        keyboard.append([
            InlineKeyboardButton("📢 مدیریت اعلان‌ها",   callback_data='admin:notif_manage'),
        ])
        keyboard.append([
            InlineKeyboardButton("📋 لاگ فعالیت",        callback_data='admin:audit_log'),
            InlineKeyboardButton("📥 خروجی اکسل",        callback_data='admin:export_excel'),
        ])
    text   = "👨‍⚕️ <b>پنل مدیریت</b>\n━━━━━━━━━━━━━━━━"
    markup = InlineKeyboardMarkup(keyboard)
    try:
        if edit and hasattr(query_or_msg, 'edit_message_text'):
            await query_or_msg.edit_message_text(text, parse_mode='HTML', reply_markup=markup)
        else:
            msg = query_or_msg if hasattr(query_or_msg, 'reply_text') else query_or_msg.message
            await msg.reply_text(text, parse_mode='HTML', reply_markup=markup)
    except Exception as e:
        logger.debug(f"_admin_menu: {e}")


async def show_admin_main(message, uid: int = None):
    await _admin_menu(message, edit=False, uid=uid)


# FIX جدید: عملیاتی که فقط ادمین ارشد (ADMIN_ID) حق انجامش را دارد —
# حتی پشتیبان/مسئول اطلاعیه/مدیر محتوای محدود هم نمی‌توانند.
ROOT_ONLY_ACTIONS = {
    'roles', 'role_add', 'role_remove', 'role_set',
    'role_add_pick', 'role_type', 'role_intake',
    'settings', 'toggle_require_sid',
    'toggle_maintenance', 'set_maintenance_text',
    'set_log_group_admin', 'set_log_group_content',
    'export_excel', 'audit_log',
    'confirm_delete_user', 'delete_user',
    'content_admins', 'ca_set', 'ca_remove',
    'notif_manage', 'notif_set_interval', 'notif_history', 'notif_retry',
    'notif_defaults', 'notif_default_toggle',
    'channel_lock', 'channel_lock_add', 'channel_lock_remove',  # FIX جدید
}


async def admin_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    uid   = update.effective_user.id
    parts  = query.data.split(':')
    action = parts[1] if len(parts) > 1 else 'main'

    # FIX جدید: سطوح دسترسی چندگانه
    if uid != ADMIN_ID:
        role_doc = await db.get_admin_role(uid)
        if not role_doc:
            await query.answer("❌ دسترسی ندارید!", show_alert=True)
            return
        if action in ROOT_ONLY_ACTIONS:
            await query.answer("❌ این بخش فقط در اختیار مدیر ارشد است.", show_alert=True)
            return
        role  = role_doc.get('role', '')
        perms = db.ROLE_PERMISSIONS.get(role, set())
        # FIX امنیتی: محدودیت دقیق هر نقش فرعی به منوی خودش —
        # غیر از 'main' (که خودش منوی فیلترشده نشان می‌دهد)، فقط
        # عمل متناسب با مجوز همان نقش اجازه دارد.
        if action == 'broadcast' and 'broadcast' not in perms:
            await query.answer("❌ شما دسترسی ارسال همگانی ندارید.", show_alert=True)
            return
        if action != 'main':
            if role == 'support':
                await query.answer(
                    "ℹ️ شما دسترسی پشتیبان دارید — از منوی «🎫 مدیریت تیکت‌ها» استفاده کنید.",
                    show_alert=True
                )
                return
            if role == 'content_scoped' and action != 'broadcast':
                await query.answer(
                    "ℹ️ شما مدیر محتوای محدود هستید — از منوی «🎓 پنل محتوا» استفاده کنید.",
                    show_alert=True
                )
                return
            if role == 'broadcaster' and action != 'broadcast':
                await query.answer(
                    "ℹ️ شما فقط دسترسی ارسال همگانی دارید.",
                    show_alert=True
                )
                return
            # FIX جدید: نقش خرخون فقط به مدیریت گزارشات دسترسی دارد
            # (که در namespace 'report:' است، نه 'admin:') — پس هر
            # اکشن admin: دیگری برایش رد می‌شود.
            if role == 'reviewer':
                await query.answer(
                    "ℹ️ شما دسترسی خرخون دارید — از منوی «⚠️ گزارشات سوال/جزوه» استفاده کنید.",
                    show_alert=True
                )
                return
            # FIX جدید: ادمین ربات فقط به کاربران، برنامه‌ها، broadcast دسترسی دارد
            if role == 'bot_admin' and action not in (
                'users', 'users_filter', 'uf_group', 'uf_intake', 'uf_clear',
                'user_detail', 'search_user', 'approve', 'reject', 'edit_group',
                'set_group', 'edit_intake', 'set_intake', 'pending', 'broadcast',
            ):
                await query.answer(
                    "ℹ️ شما دسترسی محدود دارید — کاربران، برنامه‌ها و ارسال همگانی.",
                    show_alert=True
                )
                return

    await query.answer()

    if action == 'main':
        await _admin_menu(query, uid=uid)
    elif action == 'stats':
        await _show_stats(query)

    elif action == 'bot_status':
        await _show_bot_status(query, context)

    # ══════════════════════════════════════════════
    # ⚙️ تنظیمات ربات
    # ══════════════════════════════════════════════
    elif action == 'settings':
        await _show_settings(query)

    elif action == 'toggle_require_sid':
        current = await db.get_setting('require_student_id', False)
        new_val = not current
        await db.set_setting('require_student_id', new_val)
        await query.answer("✅ شماره دانشجویی اکنون اجباری است" if new_val else "✅ شماره دانشجویی اکنون اختیاری است", show_alert=True)
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر ارشد') if admin_user else 'مدیر ارشد'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "تغییر تنظیمات ثبت‌نام", module='Settings', severity='HIGH',
            actor_role=actor_role,
            before={'شماره دانشجویی': 'اختیاری' if new_val else 'اجباری'},
            after={'شماره دانشجویی': 'اجباری' if new_val else 'اختیاری'},
            tags=['تنظیمات_ثبت_نام']
        )
        await _show_settings(query)

    elif action == 'toggle_maintenance':
        current = await db.get_setting('maintenance_mode', False)
        new_val = not current
        await db.set_setting('maintenance_mode', new_val)
        await query.answer("🔧 حالت تعمیر فعال شد" if new_val else "✅ حالت تعمیر غیرفعال شد", show_alert=True)
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر ارشد') if admin_user else 'مدیر ارشد'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "فعال‌شدن حالت تعمیر" if new_val else "غیرفعال‌شدن حالت تعمیر",
            module='Settings', severity='CRITICAL',
            actor_role=actor_role,
            before={'وضعیت': 'غیرفعال' if new_val else 'فعال'},
            after={'وضعیت': 'فعال' if new_val else 'غیرفعال'},
            tags=['حالت_تعمیر']
        )
        await _show_settings(query)

    elif action == 'set_maintenance_text':
        context.user_data['mode'] = 'set_maintenance_text'
        await query.edit_message_text(
            "🔧 <b>متن حالت تعمیر</b>\n\nمتن جدیدی که کاربران در حالت تعمیر می‌بینند را ارسال کنید:\n"
            "<i>برای بازگشت به متن پیش‌فرض، کلمه «پیشفرض» را بفرستید.</i>",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data='admin:settings')]])
        )

    elif action == 'set_log_group_admin':
        context.user_data['mode'] = 'set_log_group_admin'
        await query.edit_message_text(
            "🛡 <b>تنظیم گروه لاگ پنل ادمین</b>\n\n"
            "ربات را به گروه مورد نظر اضافه کنید، سپس یک پیام در آن گروه بفرستید و "
            "آن را به اینجا فوروارد کنید — یا مستقیماً آیدی عددی گروه (با علامت منفی) را ارسال کنید.\n\n"
            "<i>برای حذف تنظیم فعلی، کلمه «حذف» را بفرستید.</i>",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data='admin:settings')]])
        )

    elif action == 'set_log_group_content':
        context.user_data['mode'] = 'set_log_group_content'
        await query.edit_message_text(
            "🎓 <b>تنظیم گروه لاگ پنل محتوا</b>\n\n"
            "ربات را به گروه ادمین محتوا اضافه کنید، سپس یک پیام در آن گروه بفرستید و "
            "آن را به اینجا فوروارد کنید — یا مستقیماً آیدی عددی گروه (با علامت منفی) را ارسال کنید.\n\n"
            "<i>برای حذف تنظیم فعلی، کلمه «حذف» را بفرستید.</i>",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data='admin:settings')]])
        )

    elif action == 'export_excel':
        await _export_excel(query, context)

    elif action == 'channel_lock':
        await _show_channel_lock(query)

    elif action == 'channel_lock_add':
        context.user_data['mode'] = 'add_required_channel'
        await query.edit_message_text(
            "🔒 <b>افزودن کانال اجباری</b>\n\n"
            "آیدی عددی کانال (با علامت منفی، مثل <code>-1001234567890</code>) "
            "و سپس نام کانال را با کاما جدا کنید:\n\n"
            "📌 مثال:\n<code>-1001234567890, کانال اطلاع‌رسانی همیار</code>\n\n"
            "<i>⚠️ ربات باید ادمین آن کانال باشد تا بتواند عضویت را چک کند.</i>",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("❌ لغو", callback_data='admin:channel_lock')
            ]])
        )

    elif action == 'channel_lock_remove':
        ch_id = parts[2]
        await db.remove_required_channel(ch_id)
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر ارشد') if admin_user else 'مدیر ارشد'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "حذف کانال اجباری", module='Settings', severity='HIGH',
            actor_role=actor_role, target_id=ch_id, target_type='channel',
            tags=['کانال_اجباری']
        )
        await query.answer("✅ کانال حذف شد!", show_alert=True)
        await _show_channel_lock(query)

    elif action == 'audit_log':
        sev = parts[2] if len(parts) > 2 and parts[2] != 'all' else None
        await _show_audit_log(query, 'admin', sev)

    # ══════════════════════════════════════════════
    # 📢 مدیریت اعلان‌ها — FIX جدید
    # ══════════════════════════════════════════════
    elif action == 'notif_manage':
        await _show_notif_manage(query)

    elif action == 'notif_set_interval':
        hours = int(parts[2])
        await db.set_setting('resource_notif_interval_hours', hours)
        await query.answer(f"✅ فاصله اعلان منابع جدید: هر {hours} ساعت", show_alert=True)
        await _show_notif_manage(query)

    elif action == 'notif_history':
        job_name = parts[2] if len(parts) > 2 else None
        await _show_notif_history(query, job_name)

    elif action == 'notif_retry':
        run_id = parts[2]
        await _retry_failed_notif(query, context, run_id)

    elif action == 'notif_defaults':
        await _show_notif_defaults(query)

    elif action == 'notif_default_toggle':
        ntype = parts[2]
        defaults = await db.get_notif_defaults()
        new_val  = not defaults.get(ntype, True)
        await db.set_notif_default(ntype, new_val)
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر ارشد') if admin_user else 'مدیر ارشد'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "تغییر تنظیمات اعلان‌ها", module='Settings', severity='WARNING',
            actor_role=actor_role,
            details=f"پیش‌فرض {ntype}: {'روشن' if new_val else 'خاموش'}",
            tags=['تنظیمات_اعلان']
        )
        await query.answer("✅ بروزرسانی شد", show_alert=True)
        await _show_notif_defaults(query)

    # ══════════════════════════════════════════════
    # 🛡 سطوح دسترسی چندگانه ادمین (admin_roles)
    # ══════════════════════════════════════════════
    elif action == 'roles':
        # FIX باگ: پاک‌سازی state نیمه‌کاره افزودن نقش — وگرنه بعد از
        # لغو، mode='add_admin_role' باقی می‌ماند و پیام بعدی کاربر
        # در هر بخش دیگری از ربات به اشتباه به‌عنوان آیدی پارس می‌شود.
        for k in ('mode', 'new_role_type', 'new_role_intake'):
            context.user_data.pop(k, None)
        await _show_roles(query)

    elif action == 'role_add_pick':
        # انتخاب نوع نقش قبل از گرفتن آیدی
        await _show_role_type_picker(query)

    elif action == 'role_type':
        role_type = parts[2]
        context.user_data['new_role_type'] = role_type
        if role_type == 'content_scoped':
            await _show_role_intake_picker(query)
        else:
            context.user_data['mode'] = 'add_admin_role'
            await query.edit_message_text(
                f"🛡 <b>افزودن {db.ROLE_LABELS.get(role_type, role_type)}</b>\n\n"
                "آیدی عددی تلگرام کاربر را ارسال کنید:\n"
                "<i>(می‌توانید با فوروارد پیام کاربر به @userinfobot آیدی را پیدا کنید)</i>",
                parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("❌ لغو", callback_data='admin:roles')
                ]])
            )

    elif action == 'role_intake':
        intake_code = parts[2]
        context.user_data['new_role_intake'] = intake_code
        context.user_data['mode'] = 'add_admin_role'
        await query.edit_message_text(
            "🛡 <b>افزودن مدیر محتوای محدود</b>\n\n"
            "آیدی عددی تلگرام کاربر را ارسال کنید:",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("❌ لغو", callback_data='admin:roles')
            ]])
        )

    elif action == 'role_remove':
        target_uid = int(parts[2])
        await db.remove_admin_role(target_uid)
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر ارشد') if admin_user else 'مدیر ارشد'
        actor_role = await db.get_actor_role_label(uid)
        target_user_doc = await db.get_user(target_uid)
        target_label = target_user_doc.get('name', '') if target_user_doc else ''
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "حذف رول از ادمین", module='Roles', severity='CRITICAL',
            actor_role=actor_role,
            target_id=str(target_uid), target_type='user', target_label=target_label,
            tags=['نقش_ها']
        )
        await query.answer("✅ نقش حذف شد!", show_alert=True)
        await _show_roles(query)

    elif action == 'users':
        page = int(parts[2]) if len(parts) > 2 else 0
        await _show_users_list(query, page, group=context.user_data.get('filter_group'), intake=context.user_data.get('filter_intake'))
    elif action == 'users_filter':
        await _show_users_filter(query, context)
    elif action == 'uf_group':
        g = parts[2] if len(parts) > 2 and parts[2] != 'all' else None
        context.user_data['filter_group'] = g
        await _show_users_list(query, 0, group=g, intake=context.user_data.get('filter_intake'))
    elif action == 'uf_intake':
        icode = parts[2] if len(parts) > 2 and parts[2] != 'all' else None
        context.user_data['filter_intake'] = icode
        await _show_users_list(query, 0, group=context.user_data.get('filter_group'), intake=icode)
    elif action == 'uf_clear':
        context.user_data.pop('filter_group', None)
        context.user_data.pop('filter_intake', None)
        await _show_users_list(query, 0)
    elif action == 'user_detail':
        await _show_user_detail(query, context, int(parts[2]))
    elif action in ('edit_name', 'edit_group', 'edit_sid'):
        target_uid = int(parts[2])
        field_map  = {'edit_name': ('name','نام'), 'edit_group': ('group','گروه'), 'edit_sid': ('student_id','شماره دانشجویی')}
        field, label = field_map[action]
        if action == 'edit_group':
            # ویرایش گروه با دکمه — نه متن
            user_t = await db.get_user(target_uid)
            cur_g  = user_t.get('group', '') if user_t else ''
            await query.edit_message_text(
                f"👥 <b>تغییر گروه کاربر</b>\n\nگروه فعلی: <b>{cur_g or 'تعیین نشده'}</b>\n\nگروه جدید را انتخاب کنید:",
                parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup([
                    [
                        InlineKeyboardButton(f"{'✅ ' if cur_g=='1' else ''}1️⃣ گروه ۱", callback_data=f'admin:set_group:{target_uid}:1'),
                        InlineKeyboardButton(f"{'✅ ' if cur_g=='2' else ''}2️⃣ گروه ۲", callback_data=f'admin:set_group:{target_uid}:2'),
                    ],
                    [InlineKeyboardButton("🔙 بازگشت", callback_data=f'admin:user_detail:{target_uid}')],
                ])
            )
        else:
            context.user_data['edit_user'] = {'uid': target_uid, 'field': field, 'label': label}
            context.user_data['mode'] = 'edit_user'
            await query.edit_message_text(
                f"✏️ <b>ویرایش {label}</b>\n\nمقدار جدید را وارد کنید:", parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data=f'admin:user_detail:{target_uid}')]]))

    elif action == 'set_group':
        target_uid = int(parts[2])
        new_group  = parts[3] if len(parts) > 3 else '1'
        await db.update_user(target_uid, {'group': new_group})
        await query.answer(f"✅ گروه به {new_group} تغییر یافت!", show_alert=True)
        await _show_user_detail(query, context, target_uid)

    elif action == 'edit_intake':
        target_uid = int(parts[2])
        user_t     = await db.get_user(target_uid)
        cur_intake = user_t.get('intake', '') if user_t else ''
        intakes    = await db.get_all_intakes()
        keyboard   = []
        for i in intakes:
            active = cur_intake == i['code']
            keyboard.append([InlineKeyboardButton(
                f"{'✅ ' if active else ''}{i['label']}",
                callback_data=f'admin:set_intake_user:{target_uid}:{i["code"]}'
            )])
        keyboard.append([InlineKeyboardButton("❌ بدون ورودی", callback_data=f'admin:set_intake_user:{target_uid}:none')])
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data=f'admin:user_detail:{target_uid}')])
        await query.edit_message_text(
            f"📅 <b>تغییر ورودی کاربر</b>\n\nورودی فعلی: <b>{cur_intake or '—'}</b>\n\nورودی جدید را انتخاب کنید:",
            parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard)
        )

    elif action == 'set_intake_user':
        target_uid = int(parts[2])
        new_intake = '' if parts[3] == 'none' else parts[3]
        await db.update_user(target_uid, {'intake': new_intake})
        await query.answer(f"✅ ورودی به‌روز شد!", show_alert=True)
        await _show_user_detail(query, context, target_uid)
    elif action == 'suspend':
        target_uid = int(parts[2])
        target_user = await db.get_user(target_uid)
        target_name = target_user.get('name', '') if target_user else ''
        await db.update_user(target_uid, {'approved': False})
        await safe_send(context.bot, target_uid, "⚠️ دسترسی شما موقتاً تعلیق شد.")
        await query.answer("🚫 تعلیق شد!", show_alert=True)
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر') if admin_user else 'مدیر'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "مسدودسازی کاربر", module='Users', severity='HIGH',
            actor_role=actor_role,
            target_id=str(target_uid), target_type='user', target_label=target_name,
            tags=['مسدودسازی']
        )
        await _show_users_list(query, 0)
    elif action == 'approve':
        target_uid = int(parts[2])
        prev_user = await db.get_user(target_uid)
        # FIX جدید: تشخیص ثبت‌نام تازه (هنوز approved نبوده) از رفع تعلیق
        was_already_registered = bool(prev_user and prev_user.get('registered_at'))
        is_unban = was_already_registered and prev_user.get('approved') is False
        await db.update_user(target_uid, {'approved': True})
        user = await db.get_user(target_uid)
        await safe_send(context.bot, target_uid, "✅ <b>دسترسی شما تأیید شد!</b>", parse_mode='HTML', reply_markup=get_keyboard_for_uid(user, target_uid))
        await query.answer("✅ تأیید شد!", show_alert=True)
        if is_unban:
            admin_user = await db.get_user(uid)
            actor_name = admin_user.get('name', 'مدیر') if admin_user else 'مدیر'
            actor_role = await db.get_actor_role_label(uid)
            await send_audit_log(
                context.bot, 'admin', actor_name, uid,
                "رفع مسدودسازی کاربر", module='Users', severity='WARNING',
                actor_role=actor_role,
                target_id=str(target_uid), target_type='user',
                target_label=prev_user.get('name', '') if prev_user else '',
                tags=['رفع_مسدودسازی']
            )
        await _show_pending(query)
    elif action == 'reject':
        target_uid = int(parts[2])
        await db.delete_user(target_uid)
        await safe_send(context.bot, target_uid, "❌ درخواست شما رد شد.")
        await query.answer("❌ رد شد.", show_alert=True)
        await _show_pending(query)
    elif action == 'confirm_delete_user':
        target_uid = int(parts[2])
        user = await db.get_user(target_uid)
        name = user.get('name','') if user else ''
        await query.edit_message_text(
            f"⚠️ <b>حذف کاربر</b>\n\nمطمئنی می‌خواهی <b>{name}</b> را حذف کنی؟", parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("⚠️ بله، حذف", callback_data=f'admin:delete_user:{target_uid}'),
                InlineKeyboardButton("❌ لغو", callback_data=f'admin:user_detail:{target_uid}'),
            ]]))
    elif action == 'delete_user':
        target_uid = int(parts[2])
        user = await db.get_user(target_uid)
        name = user.get('name','') if user else ''
        await db.delete_user(target_uid)
        await safe_send(context.bot, target_uid, "❌ حساب شما حذف شد.")
        await query.answer(f"🗑 {name} حذف شد!", show_alert=True)
        # FIX جدید: لاگ عمل حساس — گروه ادمین
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر ارشد') if admin_user else 'مدیر ارشد'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "حذف کاربر", module='Users', severity='HIGH',
            actor_role=actor_role,
            target_id=str(target_uid), target_type='user', target_label=name,
            tags=['حذف_کاربر']
        )
        await _show_users_list(query, 0)
    elif action == 'pending':
        await _show_pending(query)
    elif action == 'search_user':
        context.user_data['mode'] = 'search_user'
        context.user_data.pop('awaiting_search', None)
        await query.edit_message_text(
            "🔍 <b>جستجوی کاربر</b>\n\nنام، شماره دانشجویی یا یوزرنیم را وارد کنید:", parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data='admin:main')]]))
    elif action == 'intakes':
        await _show_intakes(query)
    elif action == 'intake_add':
        context.user_data['mode'] = 'add_intake'
        await query.edit_message_text(
            "📅 <b>افزودن ورودی جدید</b>\n\nفرمت: <code>کد, برچسب</code>\nمثال: <code>bahman_1404, بهمن ۱۴۰۴</code>", parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data='admin:intakes')]]))
    elif action == 'intake_toggle':
        new_state = await db.toggle_intake(parts[2])
        await query.answer(f"{'✅ فعال' if new_state else '❌ غیرفعال'} شد", show_alert=True)
        await _show_intakes(query)
    elif action == 'intake_del':
        await db.delete_intake(parts[2])
        await query.answer("🗑 ورودی حذف شد!", show_alert=True)
        await _show_intakes(query)
    elif action == 'intake_view':
        code = parts[2]
        stats = await db.intake_stats(code)
        intakes = await db.get_all_intakes()
        intake = next((i for i in intakes if i['code'] == code), {})
        label = intake.get('label', code)
        groups = stats.get('groups', {})
        g_text = '\n'.join(f"  گروه {g}: {c} نفر" for g, c in groups.items()) or "  داده‌ای نیست"
        await query.edit_message_text(
            f"📅 <b>ورودی: {label}</b>\n🔑 کد: <code>{code}</code>\n━━━━━━━━━━━━━━━━\n👥 مجموع دانشجو: <b>{stats['total']}</b>\n\n<b>تفکیک گروه:</b>\n{g_text}",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به ورودی‌ها", callback_data='admin:intakes')]]))
    elif action == 'content_admins':
        admins = await db.get_content_admins()
        keyboard = []
        for a in admins:
            keyboard.append([
                InlineKeyboardButton(f"🎓 {a.get('name','')}", callback_data=f'admin:user_detail:{a["user_id"]}'),
                InlineKeyboardButton("🗑 لغو", callback_data=f'admin:ca_remove:{a["user_id"]}'),
            ])
        keyboard.append([InlineKeyboardButton("➕ دادن دسترسی", callback_data='admin:ca_grant')])
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:main')])
        await query.edit_message_text(f"🎓 <b>ادمین‌های محتوا</b> — {len(admins)} نفر", parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))
    elif action == 'ca_grant':
        users = await db.all_users(approved_only=True)
        students = [u for u in users if u.get('role','student') == 'student'][:20]
        keyboard = [[InlineKeyboardButton(f"👤 {u.get('name','')} | گروه {u.get('group','')}", callback_data=f'admin:ca_set:{u["user_id"]}')] for u in students]
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:content_admins')])
        await query.edit_message_text("➕ کاربر مورد نظر را انتخاب کنید:", reply_markup=InlineKeyboardMarkup(keyboard))
    elif action == 'ca_set':
        target_uid = int(parts[2])
        await db.update_user(target_uid, {'role': 'content_admin'})
        await safe_send(context.bot, target_uid, "🎓 <b>دسترسی ادمین محتوا به شما داده شد!</b>", parse_mode='HTML', reply_markup=content_admin_keyboard())
        await query.answer("✅ دسترسی داده شد!", show_alert=True)
        await _admin_menu(query, uid=uid)
    elif action == 'ca_remove':
        target_uid = int(parts[2])
        await db.update_user(target_uid, {'role': 'student'})
        await safe_send(context.bot, target_uid, "⚠️ دسترسی ادمین محتوای شما لغو شد.", reply_markup=main_keyboard())
        await query.answer("↩️ دسترسی لغو شد!", show_alert=True)
        await _admin_menu(query, uid=uid)
    elif action == 'qbank_manage':
        await query.edit_message_text("🧪 <b>مدیریت بانک سوال</b>", parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([
                [InlineKeyboardButton("📁 مشاهده فایل‌ها", callback_data='admin:qbank_list')],
                [InlineKeyboardButton("📤 آپلود فایل جدید", callback_data='admin:qbank_upload')],
                [InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')],
            ]))
    elif action == 'qbank_upload':
        lessons = await db.get_lessons()
        if not lessons:
            await query.edit_message_text("❌ هنوز درسی تعریف نشده.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت", callback_data='admin:qbank_manage')]]))
            return
        context.user_data['_lessons'] = lessons
        keyboard = [[InlineKeyboardButton(l, callback_data=f'admin:qbank_lesson:{i}')] for i, l in enumerate(lessons)]
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:qbank_manage')])
        await query.edit_message_text("📤 <b>آپلود بانک سوال</b>\n\nدرس را انتخاب کنید:", parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))
    elif action == 'qbank_lesson':
        idx = int(parts[2])
        lessons = context.user_data.get('_lessons', [])
        if idx < len(lessons):
            lesson = lessons[idx]
            context.user_data['qbank_lesson'] = lesson
            topics = await db.get_topics(lesson)
            context.user_data['_topics'] = topics
            keyboard = [[InlineKeyboardButton(t, callback_data=f'admin:qbank_topic:{i}')] for i, t in enumerate(topics)]
            keyboard.append([InlineKeyboardButton("📂 همه مباحث", callback_data='admin:qbank_topic:all')])
            keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:qbank_upload')])
            await query.edit_message_text(f"📚 {lesson}\n\nمبحث را انتخاب کنید:", reply_markup=InlineKeyboardMarkup(keyboard))
    elif action == 'qbank_topic':
        topics = context.user_data.get('_topics', [])
        idx = parts[2]
        topic = '' if idx == 'all' else (topics[int(idx)] if int(idx) < len(topics) else '')
        context.user_data['qbank_topic'] = topic
        context.user_data['mode'] = 'qbank_awaiting_file'
        await query.edit_message_text("📤 فایل PDF یا عکس بانک سوال را ارسال کنید:",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data='admin:qbank_manage')]]))
    elif action == 'qbank_list':
        await _show_qbank_list(query)
    elif action == 'qbank_del':
        await db.delete_qbank_file(parts[2])
        await query.answer("🗑 حذف شد!", show_alert=True)
        await _show_qbank_list(query)
    elif action == 'pending_q':
        await _pending_questions(query)
    elif action == 'approve_q':
        qid = parts[2]
        q_doc_for_log = await db.get_question_by_id(qid)
        await db.approve_question(qid)
        await query.answer("✅ تأیید شد!")
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر') if admin_user else 'مدیر'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "تأیید سوال", module='Questions', severity='INFO',
            actor_role=actor_role, target_id=qid, target_type='question',
            target_label=(q_doc_for_log.get('question', '')[:60] if q_doc_for_log else ''),
            tags=['تایید_سوال']
        )
        await _pending_questions(query)
    elif action == 'reject_q':
        qid = parts[2]
        # FIX مهم: سوال باید قبل از حذف واکشی شود تا متن آن برای
        # target_label در لاگ موجود باشد — بعد از حذف دیگر در دسترس نیست
        q_doc_for_log = await db.get_question_by_id(qid)
        await db.delete_question(qid)
        await query.answer("🗑 رد شد!")
        admin_user = await db.get_user(uid)
        actor_name = admin_user.get('name', 'مدیر') if admin_user else 'مدیر'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "رد و حذف سوال", module='Questions', severity='HIGH',
            actor_role=actor_role, target_id=qid, target_type='question',
            target_label=(q_doc_for_log.get('question', '')[:60] if q_doc_for_log else ''),
            tags=['رد_سوال']
        )
        await _pending_questions(query)

    # ══════════════════════════════════════════════
    # 📢 BROADCAST — سیستم جدید حرفه‌ای
    # ══════════════════════════════════════════════
    elif action == 'broadcast':
        await _broadcast_main(query, context)
    elif action == 'bc_target':
        target = parts[2] if len(parts) > 2 else 'all'
        context.user_data['bc_target'] = target
        await _broadcast_ask_message(query, context, target)

    elif action == 'bc_intake':
        # زیرمنوی ورودی: همه / گروه ۱ / گروه ۲
        code     = parts[2] if len(parts) > 2 else ''
        intakes  = await db.get_all_intakes()
        intake   = next((i for i in intakes if i['code'] == code), {})
        label    = intake.get('label', code)
        all_u    = await db.all_users(approved_only=True)
        all_i    = [u for u in all_u if u.get('intake') == code]
        g1_count = sum(1 for u in all_i if str(u.get('group','')) == '1')
        g2_count = sum(1 for u in all_i if str(u.get('group','')) == '2')
        keyboard = [
            [InlineKeyboardButton(
                f"👥 همه دانشجویان ورودی ({len(all_i)} نفر)",
                callback_data=f'admin:bc_target:intake_{code}'
            )],
            [
                InlineKeyboardButton(
                    f"1️⃣ گروه ۱  ({g1_count} نفر)",
                    callback_data=f'admin:bc_target:intake_{code}_g1'
                ),
                InlineKeyboardButton(
                    f"2️⃣ گروه ۲  ({g2_count} نفر)",
                    callback_data=f'admin:bc_target:intake_{code}_g2'
                ),
            ],
            [InlineKeyboardButton("🔙 بازگشت", callback_data='admin:broadcast')],
        ]
        await query.edit_message_text(
            f"📢 <b>ارسال همگانی — ورودی {label}</b>\n\n"
            f"━━━━━━━━━━━━━━━━\n"
            f"📌 گروه مورد نظر را انتخاب کنید:",
            parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard)
        )
    elif action == 'bc_cancel':
        _broadcast_clear(context)
        await query.answer("✅ لغو شد.")
        await _admin_menu(query, uid=uid)
    elif action == 'bc_confirm':
        await _broadcast_do_send(query, context)
    elif action == 'bc_schedule':
        await _broadcast_schedule_menu(query, context)
    elif action == 'bc_sched_set':
        mins = int(parts[2]) if len(parts) > 2 else 0
        context.user_data['bc_delay_min'] = mins
        await _broadcast_show_preview(query, context, scheduled=True)
    elif action == 'bc_sched_confirm':
        await _broadcast_do_send(query, context, scheduled=True)
    elif action == 'bc_edit':
        await _broadcast_ask_message(query, context, context.user_data.get('bc_target','all'), edit=True)


# ══════════════════════════════════════════════════
# 📢 توابع Broadcast
# ══════════════════════════════════════════════════

async def _broadcast_main(query, context):
    """
    ساختار صحیح:
    - همه کاربران
    - هر ورودی → زیرمنو: همه / گروه ۱ / گروه ۲
    """
    intakes   = await db.get_all_intakes()
    all_users = await db.all_users(approved_only=True)
    all_count = len(all_users)

    keyboard = [
        [InlineKeyboardButton(f"👥 همه کاربران ({all_count} نفر)", callback_data='admin:bc_target:all')],
    ]
    # هر ورودی یه دکمه جداگانه داره که زیرمنو باز میکنه
    for i in intakes:
        code   = i['code']
        label  = i['label']
        # شمارش کاربران این ورودی
        intake_users = [u for u in all_users if u.get('intake') == code]
        cnt    = len(intake_users)
        keyboard.append([InlineKeyboardButton(
            f"📅 {label} ({cnt} نفر)",
            callback_data=f'admin:bc_intake:{code}'
        )])

    keyboard.append([InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')])
    await query.edit_message_text(
        "📢 <b>ارسال همگانی</b>\n\n"
        "━━━━━━━━━━━━━━━━\n"
        "📌 <b>مرحله ۱:</b> مخاطبین پیام را انتخاب کنید:",
        parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _broadcast_ask_message(query, context, target: str, edit: bool = False):
    context.user_data['bc_target'] = target
    context.user_data['mode']      = 'broadcast'
    context.user_data.pop('bc_msg_data', None)
    target_label = _get_target_label(target)
    hint = "ویرایش" if edit else "ارسال"
    await query.edit_message_text(
        f"📢 <b>ارسال همگانی — {hint} پیام</b>\n\n"
        f"📌 مخاطب: <b>{target_label}</b>\n\n━━━━━━━━━━━━━━━━\n"
        "✍️ <b>مرحله ۲:</b> پیام خود را بنویسید:\n\n"
        "• متن (HTML پشتیبانی می‌شود)\n• عکس + کپشن\n• ویدیو + کپشن\n• فایل + کپشن\n\n"
        "<i>💡 قبل از ارسال، پیش‌نمایش نشان داده می‌شود.</i>",
        parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data='admin:bc_cancel')]]))


async def _broadcast_show_preview(query_or_msg, context, scheduled: bool = False):
    msg_data     = context.user_data.get('bc_msg_data', {})
    target       = context.user_data.get('bc_target', 'all')
    target_label = _get_target_label(target)
    delay_min    = context.user_data.get('bc_delay_min', 0)
    msg_type     = msg_data.get('type', 'text')
    type_icons   = {'text':'📝 متن','photo':'🖼 عکس','video':'🎥 ویدیو','document':'📎 فایل','voice':'🎙 ویس','audio':'🎵 صدا'}
    type_label   = type_icons.get(msg_type, '📝')
    users_list   = await _get_target_users(target)
    user_count   = len(users_list)

    if msg_type == 'text':
        text_val = msg_data.get('text','')
        content_preview = (text_val[:200] + '...') if len(text_val) > 200 else text_val
        preview_block = f"<blockquote>{content_preview}</blockquote>"
    else:
        cap = msg_data.get('caption','')
        cap_preview = (cap[:100] + '...') if cap and len(cap) > 100 else cap
        preview_block = f"[{type_label}]"
        if cap_preview:
            preview_block += f"\n<blockquote>{cap_preview}</blockquote>"

    schedule_line = ""
    if delay_min and delay_min > 0:
        h = delay_min // 60
        m = delay_min % 60
        t_str = f"{h} ساعت {m} دقیقه" if h else f"{m} دقیقه"
        send_time = (datetime.now() + timedelta(minutes=delay_min)).strftime('%H:%M')
        schedule_line = f"\n⏰ ارسال در: <b>{t_str} دیگر</b> (حدوداً ساعت {send_time})"

    info_text = (
        f"📢 <b>پیش‌نمایش و تأیید</b>\n\n"
        f"📌 مخاطب: <b>{target_label}</b>\n"
        f"👥 دریافت‌کنندگان: <b>{user_count} نفر</b>\n"
        f"📄 نوع پیام: <b>{type_label}</b>"
        f"{schedule_line}\n\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"👁 <b>پیش‌نمایش:</b>\n\n"
        f"{preview_block}\n\n"
        f"━━━━━━━━━━━━━━━━\n"
        f"آیا این پیام را ارسال می‌کنید؟"
    )

    confirm_cb = 'admin:bc_sched_confirm' if (scheduled and delay_min > 0) else 'admin:bc_confirm'
    keyboard = [
        [
            InlineKeyboardButton("✅ بله، ارسال کن", callback_data=confirm_cb),
            InlineKeyboardButton("✏️ ویرایش پیام",   callback_data='admin:bc_edit'),
        ],
        [InlineKeyboardButton("⏰ ارسال زماندار",    callback_data='admin:bc_schedule')],
        [InlineKeyboardButton("❌ لغو",               callback_data='admin:bc_cancel')],
    ]

    try:
        if hasattr(query_or_msg, 'edit_message_text'):
            await query_or_msg.edit_message_text(info_text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))
        else:
            await query_or_msg.reply_text(info_text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))
    except Exception as e:
        logger.debug(f"preview error: {e}")
        try:
            msg = query_or_msg.message if hasattr(query_or_msg, 'message') else query_or_msg
            await msg.reply_text(info_text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))
        except Exception:
            pass


async def _broadcast_schedule_menu(query, context):
    options = [
        ("⏰ ۱۵ دقیقه دیگر", 15), ("⏰ ۳۰ دقیقه دیگر", 30),
        ("⏰ ۱ ساعت دیگر",   60), ("⏰ ۲ ساعت دیگر",   120),
        ("⏰ ۶ ساعت دیگر",  360), ("⏰ ۱۲ ساعت دیگر",  720),
        ("⏰ ۲۴ ساعت دیگر",1440),
    ]
    keyboard = [[InlineKeyboardButton(label, callback_data=f'admin:bc_sched_set:{mins}')] for label, mins in options]
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به پیش‌نمایش", callback_data='admin:bc_confirm')])
    keyboard.append([InlineKeyboardButton("❌ لغو", callback_data='admin:bc_cancel')])
    await query.edit_message_text("⏰ <b>ارسال زماندار</b>\n\nچه زمانی پیام ارسال شود?",
        parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _broadcast_do_send(query, context, scheduled: bool = False):
    msg_data  = context.user_data.get('bc_msg_data', {})
    target    = context.user_data.get('bc_target', 'all')
    delay_min = context.user_data.get('bc_delay_min', 0) if scheduled else 0

    if not msg_data:
        await query.answer("❌ پیامی برای ارسال وجود ندارد!", show_alert=True)
        return

    if delay_min > 0:
        h = delay_min // 60
        m = delay_min % 60
        t_str = f"{h} ساعت {m} دقیقه" if h else f"{m} دقیقه"
        send_time = (datetime.now() + timedelta(minutes=delay_min)).strftime('%H:%M')

        context.job_queue.run_once(
            _scheduled_broadcast_job,
            when=timedelta(minutes=delay_min),
            data={'msg_data': msg_data, 'target': target, 'admin_id': ADMIN_ID},
            name=f'broadcast_{int(datetime.now().timestamp())}',
        )

        _broadcast_clear(context)
        await query.edit_message_text(
            f"✅ <b>پیام زماندار ثبت شد!</b>\n\n"
            f"⏰ ارسال خواهد شد در: <b>{t_str} دیگر</b>\n"
            f"🕐 حدوداً ساعت: <b>{send_time}</b>",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')]]))
        return

    await query.edit_message_text("⏳ <b>در حال ارسال پیام...</b>\n\nلطفاً صبر کنید.", parse_mode='HTML')
    users_list = await _get_target_users(target)
    sent, failed = await _do_broadcast_send(context.bot, users_list, msg_data)
    _broadcast_clear(context)
    # FIX طبق سند: ارسال همگانی سراسری = CRITICAL (نه WARNING)،
    # چون اگر اشتباه به همه فرستاده شود باید فوری و برجسته معلوم باشد.
    actor_id   = query.from_user.id
    actor_user = await db.get_user(actor_id)
    actor_name = actor_user.get('name', 'مدیر') if actor_user else 'مدیر'
    actor_role = await db.get_actor_role_label(actor_id)
    await send_audit_log(
        context.bot, 'admin', actor_name, actor_id,
        "ارسال همگانی", module='Notifications', severity='CRITICAL',
        actor_role=actor_role,
        target_type='broadcast', target_label=f"هدف: {target}",
        details=f"موفق: {sent} نفر | ناموفق: {failed} نفر",
        tags=['ارسال_همگانی']
    )
    await query.edit_message_text(
        f"📢 <b>ارسال همگانی تمام شد</b>\n\n━━━━━━━━━━━━━━━━\n"
        f"✅ موفق: <b>{sent} نفر</b>\n❌ ناموفق: <b>{failed} نفر</b>\n📊 مجموع: <b>{sent+failed} نفر</b>",
        parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')]]))


async def _scheduled_broadcast_job(context: ContextTypes.DEFAULT_TYPE):
    data = context.job.data
    msg_data = data.get('msg_data', {})
    target   = data.get('target', 'all')
    admin_id = data.get('admin_id', ADMIN_ID)
    users_list = await _get_target_users(target)
    sent, failed = await _do_broadcast_send(context.bot, users_list, msg_data)
    try:
        await context.bot.send_message(admin_id,
            f"📢 <b>ارسال زماندار انجام شد</b>\n\n✅ موفق: <b>{sent} نفر</b>\n❌ ناموفق: <b>{failed} نفر</b>",
            parse_mode='HTML')
    except Exception:
        pass


async def _do_broadcast_send(bot, users_list: list, msg_data: dict) -> tuple:
    sent, failed = 0, 0
    msg_type = msg_data.get('type', 'text')
    caption  = msg_data.get('caption', '')
    file_id  = msg_data.get('file_id', '')
    text_val = msg_data.get('text', '')

    for i, u in enumerate(users_list):
        uid = u['user_id']
        try:
            if msg_type == 'text':
                await bot.send_message(uid, text_val, parse_mode='HTML')
            elif msg_type == 'photo':
                await bot.send_photo(uid, file_id, caption=caption, parse_mode='HTML')
            elif msg_type == 'video':
                await bot.send_video(uid, file_id, caption=caption, parse_mode='HTML')
            elif msg_type == 'document':
                await bot.send_document(uid, file_id, caption=caption, parse_mode='HTML')
            elif msg_type == 'voice':
                await bot.send_voice(uid, file_id, caption=caption)
            elif msg_type == 'audio':
                await bot.send_audio(uid, file_id, caption=caption, parse_mode='HTML')
            sent += 1
        except Exception:
            failed += 1
        if i % 30 == 29:
            await asyncio.sleep(1)
        else:
            await asyncio.sleep(0.05)

    return sent, failed


async def _get_target_users(target: str) -> list:
    """
    فرمت‌های پشتیبانی‌شده:
      all              → همه کاربران
      g1 / g2          → گروه ۱ یا ۲ از همه ورودی‌ها
      intake_CODE      → همه کاربران یک ورودی
      intake_CODE_g1   → گروه ۱ از ورودی CODE
      intake_CODE_g2   → گروه ۲ از ورودی CODE
    """
    all_users = await db.all_users(approved_only=True)
    if target == "all":
        return all_users
    elif target == "g1":
        return [u for u in all_users if str(u.get("group", "")) == "1"]
    elif target == "g2":
        return [u for u in all_users if str(u.get("group", "")) == "2"]
    elif target.startswith("intake_"):
        rest = target[7:]
        if rest.endswith("_g1"):
            code = rest[:-3]
            return [u for u in all_users
                    if u.get("intake") == code and str(u.get("group","")) == "1"]
        elif rest.endswith("_g2"):
            code = rest[:-3]
            return [u for u in all_users
                    if u.get("intake") == code and str(u.get("group","")) == "2"]
        else:
            return [u for u in all_users if u.get("intake") == rest]
    return all_users


def _get_target_label(target: str) -> str:
    labels = {'all': 'همه کاربران', 'g1': 'گروه ۱ (همه ورودی‌ها)', 'g2': 'گروه ۲ (همه ورودی‌ها)'}
    if target in labels:
        return labels[target]
    if target.startswith('intake_'):
        rest = target[7:]
        if rest.endswith('_g1'):
            return f"ورودی {rest[:-3]} — گروه ۱"
        elif rest.endswith('_g2'):
            return f"ورودی {rest[:-3]} — گروه ۲"
        return f"ورودی {rest}"
    return target


def _broadcast_clear(context):
    for key in ['bc_target', 'bc_msg_data', 'bc_delay_min', 'mode']:
        context.user_data.pop(key, None)


async def admin_broadcast_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """دریافت پیام و نمایش preview — فراخوانی از unified handlers در bot.py"""
    uid = update.effective_user.id
    if uid != ADMIN_ID:
        return
    if context.user_data.get('mode') != 'broadcast':
        return

    msg = update.message
    if msg.text:
        context.user_data['bc_msg_data'] = {'type': 'text', 'text': msg.text}
    elif msg.photo:
        context.user_data['bc_msg_data'] = {'type': 'photo', 'file_id': msg.photo[-1].file_id, 'caption': msg.caption or ''}
    elif msg.video:
        context.user_data['bc_msg_data'] = {'type': 'video', 'file_id': msg.video.file_id, 'caption': msg.caption or ''}
    elif msg.document:
        context.user_data['bc_msg_data'] = {'type': 'document', 'file_id': msg.document.file_id, 'caption': msg.caption or ''}
    elif msg.voice:
        context.user_data['bc_msg_data'] = {'type': 'voice', 'file_id': msg.voice.file_id, 'caption': ''}
    elif msg.audio:
        context.user_data['bc_msg_data'] = {'type': 'audio', 'file_id': msg.audio.file_id, 'caption': msg.caption or ''}
    else:
        await msg.reply_text("❌ این نوع پیام پشتیبانی نمی‌شود.\nلطفاً متن، عکس، ویدیو یا فایل ارسال کنید.")
        return

    context.user_data['mode'] = ''  # reset mode
    await _broadcast_show_preview(msg, context)


# ══════════════════════════════════════════════════
# نمایش‌دهنده‌ها
# ══════════════════════════════════════════════════

async def _show_bot_status(query, context):
    """
    FIX جدید: مانیتورینگ واقعی سلامت سیستم با psutil — مصرف RAM/CPU
    پروسه و کانتینر، uptime واقعی، تعداد کاربران آنلاین تقریبی.
    روی Railway این اعداد مربوط به همان کانتینر ربات است (نه کل
    سرور فیزیکی) — دقیقاً همان چیزی که برای پایش خود ربات لازم است.
    """
    import time
    from datetime import datetime

    db_status = "disconnected"
    db_ping   = "—"
    try:
        t0 = time.monotonic()
        await db.client.admin.command("ping")
        db_ping   = f"{int((time.monotonic()-t0)*1000)} ms"
        db_status = "✅ متصل"
    except Exception as e:
        db_status = f"❌ خطا: {str(e)[:30]}"

    jobs_info = []
    try:
        if context.application.job_queue:
            for job in context.application.job_queue.jobs():
                nxt = job.next_t
                nxt_str = nxt.strftime("%H:%M") if nxt else "—"
                jobs_info.append(f"  ⏰ {job.name}  |  بعدی: {nxt_str}")
    except Exception:
        pass
    jobs_text = "\n".join(jobs_info) if jobs_info else "  —"

    # FIX جدید: متریک‌های واقعی سیستم با psutil
    sys_lines = []
    try:
        import psutil, os
        proc = psutil.Process(os.getpid())

        # حافظه پروسه ربات (مهم‌ترین عدد — واقعاً ربات چقدر مصرف می‌کند)
        mem_mb = proc.memory_info().rss / 1024 / 1024

        # حافظه کل کانتینر
        vm = psutil.virtual_memory()
        vm_used_mb  = vm.used / 1024 / 1024
        vm_total_mb = vm.total / 1024 / 1024

        # CPU (۰.۳ ثانیه نمونه‌گیری — سریع و کافی برای یک عدد لحظه‌ای)
        cpu_pct = psutil.cpu_percent(interval=0.3)

        # uptime واقعی پروسه ربات (نه زمان سرور)
        uptime_sec = time.time() - proc.create_time()
        h, rem = divmod(int(uptime_sec), 3600)
        m, s_  = divmod(rem, 60)
        uptime_str = f"{h} ساعت {m} دقیقه" if h else f"{m} دقیقه {s_} ثانیه"

        # رنگ‌بندی هشدار بر اساس فشار منابع
        ram_icon = "🟢" if vm.percent < 70 else "🟡" if vm.percent < 90 else "🔴"
        cpu_icon = "🟢" if cpu_pct < 70 else "🟡" if cpu_pct < 90 else "🔴"

        sys_lines = [
            "",
            "━━━━━━━━━━━━━━━━",
            "🖥 <b>سلامت سیستم</b>",
            "",
            f"{ram_icon} <b>RAM کانتینر:</b> {vm_used_mb:.0f} / {vm_total_mb:.0f} MB  ({vm.percent}%)",
            f"   └ مصرف خود ربات: {mem_mb:.1f} MB",
            f"{cpu_icon} <b>CPU:</b> {cpu_pct}%",
            f"⏱ <b>مدت کارکرد ربات:</b> {uptime_str}",
        ]
    except ImportError:
        sys_lines = [
            "",
            "━━━━━━━━━━━━━━━━",
            "🖥 <b>سلامت سیستم:</b> ⚠️ psutil نصب نیست",
        ]
    except Exception as e:
        sys_lines = [
            "",
            "━━━━━━━━━━━━━━━━",
            f"🖥 <b>سلامت سیستم:</b> ⚠️ خطا در خوانش ({str(e)[:40]})",
        ]

    # FIX جدید: کاربران آنلاین تقریبی (فعالیت در ۳۰ دقیقه و امروز)
    try:
        online_30m = await db.count_active_users(30)
        active_today = await db.count_active_users_today()
        online_line = f"🟢 آنلاین (۳۰ دقیقه اخیر): <b>{online_30m}</b>  |  فعال امروز: <b>{active_today}</b>"
    except Exception:
        online_line = "🟢 آنلاین: داده در دسترس نیست"

    s = await db.global_stats()
    now_str = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    lines_t = [
        "📡 <b>وضعیت ربات</b>",
        "━━━━━━━━━━━━━━━━",
        "",
        f"🗄 <b>دیتابیس:</b> {db_status}",
        f"🏓 <b>پینگ DB:</b> {db_ping}",
        "",
        "⏰ <b>Job های فعال:</b>",
        jobs_text,
    ] + sys_lines + [
        "",
        "━━━━━━━━━━━━━━━━",
        "📊 <b>آمار کلی</b>",
        "",
        online_line,
        f"👥 کاربران تأیید: <b>{s['users']}</b>",
        f"⏳ منتظر تأیید: <b>{s['pending']}</b>",
        f"🧪 سوال تأییدشده: <b>{s['questions']}</b>",
        f"📁 محتوای علوم پایه: <b>{s.get('bs_content', 0)}</b>",
        f"📖 رفرنس ها: <b>{s.get('ref_files', 0)}</b>",
        f"🎫 تیکت باز: <b>{s.get('open_tickets', 0)}</b>",
        "",
        "━━━━━━━━━━━━━━━━",
        f"🕐 زمان سرور: <code>{now_str}</code>",
    ]
    text = "\n".join(lines_t)
    await query.edit_message_text(
        text, parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 بروزرسانی", callback_data="admin:bot_status")],
            [InlineKeyboardButton("🔙 بازگشت به پنل", callback_data="admin:main")],
        ])
    )


async def _show_stats(query):
    s = await db.global_stats()
    text = (
        "📊 <b>آمار سیستم</b>\n━━━━━━━━━━━━━━━━\n\n"
        f"👥 کاربران تأیید: <b>{s['users']}</b>  |  ⏳ منتظر: <b>{s['pending']}</b>\n"
        f"🆕 جدید این هفته: <b>{s.get('new_users_week',0)}</b>\n"
        f"🎓 ادمین محتوا: <b>{s.get('content_admins',0)}</b>\n\n"
        f"🔬 <b>علوم پایه:</b>\n"
        f"  📖 درس: <b>{s.get('bs_lessons',0)}</b>  📌 جلسه: <b>{s.get('bs_sessions',0)}</b>  📁 فایل: <b>{s.get('bs_content',0)}</b>\n\n"
        f"📚 <b>رفرنس‌ها:</b>\n"
        f"  📖 درس: <b>{s.get('ref_subjects',0)}</b>  📘 کتاب: <b>{s.get('ref_books',0)}</b>\n\n"
        f"🧪 بانک سوال: <b>{s['questions']}</b>  📁 فایل: <b>{s.get('qbank_files',0)}</b>\n"
        f"🎫 تیکت‌های باز: <b>{s.get('open_tickets',0)}</b>"
    )
    await query.edit_message_text(text, parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("🔄 بروزرسانی", callback_data='admin:stats')],
            [InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')],
        ]))


async def _show_users_list(query, page: int = 0, group: str = None, intake: str = None):
    all_users = await db.all_users(approved_only=False)
    if group:
        all_users = [u for u in all_users if u.get('group') == group]
    if intake:
        all_users = [u for u in all_users if u.get('intake') == intake]
    per_page = 8
    total    = len(all_users)
    approved = sum(1 for u in all_users if u.get('approved'))
    start    = page * per_page
    chunk    = all_users[start:start + per_page]
    filter_parts = []
    if group:  filter_parts.append(f"گروه {group}")
    if intake: filter_parts.append(f"ورودی {intake}")
    filter_label = f" | 🔽 {' + '.join(filter_parts)}" if filter_parts else ""
    text = f"👥 <b>کاربران{filter_label}</b>\n✅ تأیید: {approved} | ⏳ منتظر: {total-approved} | مجموع: {total}\n\n"
    keyboard = []
    for u in chunk:
        icon  = "✅" if u.get('approved') else "⏳"
        role  = "🎓" if u.get('role') == 'content_admin' else ""
        itak  = f" | {u.get('intake','')}" if u.get('intake') else ""
        keyboard.append([InlineKeyboardButton(
            f"{icon}{role} {u.get('name','')[:10]} | گروه {u.get('group','')}{itak}",
            callback_data=f'admin:user_detail:{u["user_id"]}')])
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️ قبلی", callback_data=f'admin:users:{page-1}'))
    if start + per_page < total:
        nav.append(InlineKeyboardButton("بعدی ▶️", callback_data=f'admin:users:{page+1}'))
    if nav:
        keyboard.append(nav)
    keyboard.append([InlineKeyboardButton("🔽 فیلتر", callback_data='admin:users_filter'), InlineKeyboardButton("🔍 جستجو", callback_data='admin:search_user')])
    if group or intake:
        keyboard.append([InlineKeyboardButton("❌ حذف فیلتر", callback_data='admin:uf_clear')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')])
    await query.edit_message_text(text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_users_filter(query, context):
    intakes  = await db.get_all_intakes()
    f_group  = context.user_data.get('filter_group')
    f_intake = context.user_data.get('filter_intake')
    keyboard = [
        [InlineKeyboardButton("━━ فیلتر گروه ━━", callback_data='admin:main')],
        [
            InlineKeyboardButton(f"{'✅' if not f_group else '⬜'} همه", callback_data='admin:uf_group:all'),
            InlineKeyboardButton(f"{'✅' if f_group=='1' else '⬜'} گروه ۱", callback_data='admin:uf_group:1'),
            InlineKeyboardButton(f"{'✅' if f_group=='2' else '⬜'} گروه ۲", callback_data='admin:uf_group:2'),
        ],
        [InlineKeyboardButton("━━ فیلتر ورودی ━━", callback_data='admin:main')],
        [InlineKeyboardButton("همه ورودی‌ها", callback_data='admin:uf_intake:all')],
    ]
    for i in intakes:
        active = f_intake == i['code']
        keyboard.append([InlineKeyboardButton(f"{'✅' if active else '⬜'} {i['label']}", callback_data=f'admin:uf_intake:{i["code"]}')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:users:0')])
    active_filters = []
    if f_group:  active_filters.append(f"گروه {f_group}")
    if f_intake: active_filters.append(f_intake)
    current = f"فعال: {' + '.join(active_filters)}" if active_filters else "بدون فیلتر"
    await query.edit_message_text(f"🔽 <b>فیلتر کاربران</b>\n{current}", parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_user_detail(query, context, target_uid: int):
    user = await db.get_user(target_uid)
    if not user:
        await query.answer("کاربر پیدا نشد!", show_alert=True)
        return
    stats   = await db.user_stats(target_uid)
    status  = "✅ تأیید شده" if user.get('approved') else "⏳ در انتظار"
    role_m  = {'student': '🧑‍🎓 دانشجو', 'content_admin': '🎓 ادمین محتوا'}
    role_t  = role_m.get(user.get('role','student'), user.get('role',''))
    uname   = f"@{user['username']}" if user.get('username') else 'ندارد'
    tickets = await db.ticket_get_user(target_uid)
    open_t  = sum(1 for t in tickets if t['status'] == 'open')
    text = (
        f"👤 <b>پروفایل کاربر</b>\n━━━━━━━━━━━━━━━━\n\n"
        f"📛 نام: <b>{user.get('name','')}</b>\n"
        f"🎓 شماره: <code>{user.get('student_id','') or '—'}</code>\n"
        f"👥 گروه: <b>{user.get('group','')}</b>\n"
        f"📱 یوزرنیم: {uname}\n"
        f"🆔 آیدی: <code>{target_uid}</code>\n"
        f"🔘 وضعیت: {status}  |  نقش: {role_t}\n"
        f"📅 ورودی: <b>{user.get('intake','') or 'ثبت نشده'}</b>\n"
        f"📅 ثبت‌نام: {user.get('registered_at','')[:10]}\n\n"
        f"📊 <b>آمار:</b>\n"
        f"  📥 دانلود: {stats['downloads']}  🧪 سوال: {stats['total_answers']}  ✅ صحیح: {stats['correct_answers']}\n"
        f"  📈 درصد: {stats['percentage']}%  🔥 هفتگی: {stats['week_activity']}\n"
        f"  🎫 تیکت باز: {open_t}"
    )
    keyboard = [
        [
            InlineKeyboardButton("✏️ ویرایش نام",  callback_data=f'admin:edit_name:{target_uid}'),
            InlineKeyboardButton("✏️ ویرایش گروه", callback_data=f'admin:edit_group:{target_uid}'),
        ],
        [InlineKeyboardButton("📅 ویرایش ورودی", callback_data=f'admin:edit_intake:{target_uid}')],
    ]
    if user.get('role','student') == 'student':
        keyboard.append([InlineKeyboardButton("🎓 دادن دسترسی محتوا", callback_data=f'admin:ca_set:{target_uid}')])
    elif user.get('role') == 'content_admin':
        keyboard.append([InlineKeyboardButton("↩️ لغو دسترسی محتوا", callback_data=f'admin:ca_remove:{target_uid}')])
    if user.get('approved'):
        keyboard.append([InlineKeyboardButton("🚫 تعلیق", callback_data=f'admin:suspend:{target_uid}')])
    else:
        keyboard.append([InlineKeyboardButton("✅ تأیید", callback_data=f'admin:approve:{target_uid}'), InlineKeyboardButton("❌ رد", callback_data=f'admin:reject:{target_uid}')])
    keyboard.append([InlineKeyboardButton("🗑 حذف کامل", callback_data=f'admin:confirm_delete_user:{target_uid}')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به لیست", callback_data='admin:users:0')])
    await query.edit_message_text(text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_pending(query):
    pending = await db.pending_users()
    if not pending:
        await query.edit_message_text("✅ هیچ کاربر در انتظاری وجود ندارد.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')]]))
        return
    keyboard = []
    for u in pending:
        uid = u['user_id']
        keyboard.append([InlineKeyboardButton(f"👤 {u.get('name','')} | {u.get('student_id','') or 'بدون شماره'} | گروه {u.get('group','')}", callback_data=f'admin:user_detail:{uid}')])
        keyboard.append([InlineKeyboardButton("✅ تأیید", callback_data=f'admin:approve:{uid}'), InlineKeyboardButton("❌ رد", callback_data=f'admin:reject:{uid}')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')])
    await query.edit_message_text(f"⏳ <b>کاربران در انتظار</b> — {len(pending)} نفر", parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _pending_questions(query):
    questions = await db.pending_questions()
    if not questions:
        await query.edit_message_text("✅ هیچ سوال در انتظاری وجود ندارد.",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')]]))
        return
    keyboard = []
    for q in questions[:8]:
        qid    = str(q['_id'])
        opts   = q.get('options', [])
        ltrs   = ['الف', 'ب', 'ج', 'د']
        correct_idx = q.get('correct_answer', 0)
        correct_txt = opts[correct_idx] if correct_idx < len(opts) else '—'
        diff_map = {'easy':'🟢 آسان','medium':'🟡 متوسط','hard':'🔴 سخت'}
        diff_txt = diff_map.get(q.get('difficulty',''), q.get('difficulty',''))
        keyboard.append([InlineKeyboardButton(f"📋 {q.get('lesson','')} | {q.get('topic','')} | {diff_txt}", callback_data='admin:pending_q')])
        keyboard.append([InlineKeyboardButton(f"❓ {q.get('question','')[:50]}", callback_data='admin:pending_q')])
        opts_short = ' | '.join(f"{ltrs[i]}) {o[:15]}" for i, o in enumerate(opts[:4]))
        keyboard.append([InlineKeyboardButton(f"گزینه‌ها: {opts_short[:50]}", callback_data='admin:pending_q')])
        keyboard.append([InlineKeyboardButton(f"✅ جواب: {correct_txt[:20]}", callback_data='admin:pending_q')])
        keyboard.append([
            InlineKeyboardButton("✅ تأیید", callback_data=f'admin:approve_q:{qid}'),
            InlineKeyboardButton("🗑 رد",    callback_data=f'admin:reject_q:{qid}'),
        ])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')])
    await query.edit_message_text(
        f"⏳ <b>سوالات در انتظار تأیید</b> — {len(questions)} سوال\n━━━━━━━━━━━━━━━━",
        parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_qbank_list(query):
    files = await db.get_qbank_files()
    keyboard = []
    for f in files[:15]:
        fid = str(f['_id'])
        keyboard.append([
            InlineKeyboardButton(f"📁 {f.get('lesson','')} — {f.get('topic','')[:15]}", callback_data='admin:qbank_list'),
            InlineKeyboardButton("🗑", callback_data=f'admin:qbank_del:{fid}'),
        ])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:qbank_manage')])
    await query.edit_message_text(
        f"📁 <b>فایل‌های بانک سوال</b> — {len(files)} فایل" if files else "❌ فایلی آپلود نشده.",
        parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_notif_manage(query):
    """
    FIX جدید: مدیریت اعلان‌ها از پنل ادمین — تنظیم فاصله زمانی
    اعلان منابع جدید + دسترسی به تاریخچه ارسال هر job.
    """
    interval = await db.get_setting('resource_notif_interval_hours', 24)
    pending  = await db.get_unnotified_resources()

    text = (
        "📢 <b>مدیریت اعلان‌ها</b>\n━━━━━━━━━━━━━━━━\n\n"
        f"📚 <b>فاصله اعلان منابع جدید:</b> هر {interval} ساعت\n"
        f"⏳ منابع در انتظار اعلام: <b>{len(pending)}</b> مورد\n\n"
        "━━━━━━━━━━━━━━━━\n"
        "برای مشاهده تاریخچه ارسال هر دسته از اعلان‌ها، یکی را انتخاب کنید:"
    )
    keyboard = [
        [
            InlineKeyboardButton("24 ساعت" + (" ✅" if interval == 24 else ""), callback_data='admin:notif_set_interval:24'),
            InlineKeyboardButton("48 ساعت" + (" ✅" if interval == 48 else ""), callback_data='admin:notif_set_interval:48'),
            InlineKeyboardButton("72 ساعت" + (" ✅" if interval == 72 else ""), callback_data='admin:notif_set_interval:72'),
        ],
        [InlineKeyboardButton("⚙️ وضعیت پیش‌فرض اعلان‌ها (کاربر جدید)", callback_data='admin:notif_defaults')],
        [InlineKeyboardButton("📚 تاریخچه: منابع جدید",   callback_data='admin:notif_history:new_resources')],
        [InlineKeyboardButton("📝 تاریخچه: یادآوری امتحان", callback_data='admin:notif_history:exam_reminder')],
        [InlineKeyboardButton("🧪 تاریخچه: سوال روزانه",   callback_data='admin:notif_history:daily_question')],
        [InlineKeyboardButton("📋 همه اجراهای اخیر",        callback_data='admin:notif_history')],
        [InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')],
    ]
    await query.edit_message_text(text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_notif_defaults(query):
    """
    FIX طبق سند: تعیین وضعیت پیش‌فرض (روشن/خاموش) هر دسته اعلان
    برای کاربران تازه ثبت‌نام‌شده — قابل تغییر، بدون اثر روی
    کاربران فعلی که خودشان تنظیم کرده‌اند.
    """
    from notifications import NOTIF_ITEMS
    defaults = await db.get_notif_defaults()
    text = (
        "⚙️ <b>وضعیت پیش‌فرض اعلان‌ها</b>\n━━━━━━━━━━━━━━━━\n\n"
        "این تنظیمات فقط روی <b>کاربران جدید</b> اعمال می‌شود؛ "
        "کاربران فعلی همان انتخاب خودشان را دارند."
    )
    keyboard = []
    for key, label, _ in NOTIF_ITEMS:
        is_on = defaults.get(key, True)
        icon  = "🔔" if is_on else "🔕"
        keyboard.append([InlineKeyboardButton(
            f"{icon} {label} — {'روشن' if is_on else 'خاموش'}",
            callback_data=f'admin:notif_default_toggle:{key}'
        )])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:notif_manage')])
    await query.edit_message_text(text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_notif_history(query, job_name: str = None):
    """
    FIX جدید: نمایش تاریخچه اجراهای notif_runs — وضعیت موفق/ناموفق
    هر اجرا، با دکمه retry برای اجراهایی که شکست داشته‌اند.
    """
    runs = await db.get_recent_notif_runs(job_name, limit=10)
    job_label = {
        'new_resources': '📚 منابع جدید', 'exam_reminder': '📝 یادآوری امتحان',
        'daily_question': '🧪 سوال روزانه',
    }.get(job_name, '📋 همه اعلان‌ها')

    if not runs:
        text = f"{job_label}\n\nهنوز هیچ اجرایی ثبت نشده."
        keyboard = [[InlineKeyboardButton("🔙 بازگشت", callback_data='admin:notif_manage')]]
    else:
        lines = [f"{job_label}\n━━━━━━━━━━━━━━━━"]
        keyboard = []
        for r in runs:
            started = r.get('started_at', '')[:16].replace('T', ' ')
            status  = r.get('status', 'running')
            icon    = {'completed': '✅', 'error': '❌', 'skipped': '⏭', 'running': '⏳'}.get(status, '❔')
            sent    = r.get('sent', 0)
            failed  = r.get('failed', 0)
            lines.append(
                f"\n{icon} <code>{started}</code>\n"
                f"   ✅ موفق: {sent}  |  ❌ ناموفق: {failed}"
            )
            if failed > 0:
                rid = str(r['_id'])
                keyboard.append([InlineKeyboardButton(
                    f"🔄 تلاش مجدد — {started}", callback_data=f'admin:notif_retry:{rid}'
                )])
        text = '\n'.join(lines)
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:notif_manage')])

    await query.edit_message_text(text[:4000], parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _retry_failed_notif(query, context, run_id: str):
    """
    FIX جدید: ارسال مجدد برای کاربرانی که در یک اجرای قبلی fail شدند.
    """
    failed_ids = await db.get_failed_notif_targets(run_id)
    if not failed_ids:
        await query.answer("✅ موردی برای تلاش مجدد نیست.", show_alert=True)
        return
    sent, failed = 0, 0
    for uid_target in failed_ids:
        try:
            await context.bot.send_message(
                uid_target,
                "🔔 <b>یادآوری</b>\n\nاین پیام به دلیل خطای موقت دوباره ارسال شد. "
                "برای جزئیات به بخش‌های مربوطه ربات مراجعه کنید.",
                parse_mode='HTML'
            )
            sent += 1
        except Exception:
            failed += 1
    await query.answer(f"✅ {sent} ارسال موفق، {failed} هنوز ناموفق", show_alert=True)


async def _show_channel_lock(query):
    """
    FIX جدید: مدیریت قفل اجباری عضویت کانال — لیست کانال‌های فعلی
    + دکمه افزودن/حذف. اگر هیچ کانالی نباشد، قفل عملاً غیرفعال است.
    """
    channels = await db.get_required_channels()
    text = (
        "🔒 <b>قفل اجباری عضویت کانال</b>\n━━━━━━━━━━━━━━━━\n\n"
        + (
            f"📌 {len(channels)} کانال فعال — کاربران عادی باید عضو همه آن‌ها باشند:\n\n"
            if channels else
            "⬜ هیچ کانالی تنظیم نشده — قفل غیرفعال است.\n\n"
        )
    )
    keyboard = []
    for ch in channels:
        keyboard.append([InlineKeyboardButton(
            f"🗑 {ch['title']}", callback_data=f'admin:channel_lock_remove:{ch["id"]}'
        )])
    keyboard.append([InlineKeyboardButton("➕ افزودن کانال جدید", callback_data='admin:channel_lock_add')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به تنظیمات", callback_data='admin:settings')])
    await query.edit_message_text(text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_settings(query):
    """
    FIX جدید: صفحه تنظیمات کلی ربات — شامل اجباری‌بودن شماره
    دانشجویی، حالت تعمیر و نگهداری، و گروه‌های لاگ تلگرامی.
    """
    require_sid = await db.get_setting('require_student_id', False)
    maint       = await db.get_setting('maintenance_mode', False)
    log_admin   = await db.get_setting('log_group_admin', None)
    log_content = await db.get_setting('log_group_content', None)
    missing     = await db.users_missing_student_id()

    sid_status   = "✅ فعال (اجباری)" if require_sid else "⬜ غیرفعال (اختیاری)"
    sid_toggle    = "🔴 غیرفعال کردن" if require_sid else "🟢 فعال کردن"
    maint_status = "🔧 فعال (ربات در دسترس کاربران عادی نیست)" if maint else "✅ غیرفعال (ربات عادی کار می‌کند)"
    maint_toggle  = "🟢 خاموش کردن حالت تعمیر" if maint else "🔴 روشن کردن حالت تعمیر"
    admin_grp_txt = f"<code>{log_admin}</code>" if log_admin else "تنظیم نشده"
    content_grp_txt = f"<code>{log_content}</code>" if log_content else "تنظیم نشده"

    text = (
        "⚙️ <b>تنظیمات ربات</b>\n━━━━━━━━━━━━━━━━\n\n"
        f"🎓 <b>الزامی بودن شماره دانشجویی:</b> {sid_status}\n"
        f"👥 کاربران بدون شماره دانشجویی: <b>{len(missing)}</b> نفر\n\n"
        f"🔧 <b>حالت تعمیر و نگهداری:</b> {maint_status}\n\n"
        f"🛡 گروه لاگ پنل ادمین: {admin_grp_txt}\n"
        f"🎓 گروه لاگ پنل محتوا: {content_grp_txt}\n"
    )
    channels = await db.get_required_channels()
    channel_label = f"🔒 قفل کانال: {len(channels)} کانال فعال" if channels else "🔓 قفل کانال: غیرفعال"

    keyboard = [
        [InlineKeyboardButton(sid_toggle, callback_data='admin:toggle_require_sid')],
        [InlineKeyboardButton(maint_toggle, callback_data='admin:toggle_maintenance')],
        [InlineKeyboardButton("✏️ متن حالت تعمیر", callback_data='admin:set_maintenance_text')],
        [InlineKeyboardButton("🛡 تنظیم گروه لاگ ادمین", callback_data='admin:set_log_group_admin')],
        [InlineKeyboardButton("🎓 تنظیم گروه لاگ محتوا", callback_data='admin:set_log_group_content')],
        [InlineKeyboardButton(channel_label, callback_data='admin:channel_lock')],
        [InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')],
    ]
    await query.edit_message_text(text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_audit_log(query, category: str, min_severity: str = None, module: str = None):
    """
    FIX بازطراحی کامل طبق سند جدید Audit Log — نمایش هر لاگ با ساختار
    actor{name,role}/target{label}/changes/تگ‌ها، هماهنگ با مدل داده
    جدید database.py. در کمتر از ۵ ثانیه باید همه‌چیز معلوم باشد.
    """
    logs = await db.get_recent_logs(category, min_severity, module, limit=20)
    sev_icon = {'INFO': '🟢', 'WARNING': '🟡', 'HIGH': '🟠', 'CRITICAL': '🔴'}

    if not logs:
        text = "📋 <b>لاگ فعالیت</b>\n\nهیچ موردی با این فیلتر یافت نشد."
    else:
        lines = ["📋 <b>لاگ فعالیت</b>\n━━━━━━━━━━━━━━━━"]
        for log in logs:
            ts = log.get('timestamp', '')[:16].replace('T', ' ')
            icon = sev_icon.get(log.get('severity', 'INFO'), '🟢')
            module_en = log.get('module', '')
            module_fa = db.MODULE_LABELS_FA.get(module_en, module_en)
            module_tag = f" [{module_fa}]" if module_fa else ""

            actor = log.get('actor', {})
            actor_name = actor.get('name', '') if isinstance(actor, dict) else log.get('actor_name', '')
            actor_role = actor.get('role', '') if isinstance(actor, dict) else log.get('actor_role', '')

            target = log.get('target', {})
            target_label = target.get('label', '') if isinstance(target, dict) else ''

            lines.append(f"\n{icon} <code>{ts}</code>{module_tag}")
            lines.append(f"👤 {actor_name} ({actor_role}) — <b>{log.get('action','')}</b>")
            if target_label:
                lines.append(f"🎯 {target_label}")

            changes = log.get('changes', [])
            if changes:
                for ch in changes:
                    lines.append(f"   {ch.get('field','')}: {ch.get('before','—')} ← {ch.get('after','—')}")
            elif log.get('details'):
                lines.append(f"   📝 {log['details']}")
        text = '\n'.join(lines)

    filter_label = {
        None: "همه سطوح", 'WARNING': "🟡 WARNING به بالا",
        'HIGH': "🟠 HIGH به بالا", 'CRITICAL': "🔴 فقط CRITICAL",
    }.get(min_severity, "همه سطوح")
    text = f"🔽 فیلتر: {filter_label}\n\n" + text

    keyboard = [
        [
            InlineKeyboardButton("همه" + (" ✅" if not min_severity else ""), callback_data='admin:audit_log:all'),
            InlineKeyboardButton("🟡 WARNING+" + (" ✅" if min_severity == 'WARNING' else ""), callback_data='admin:audit_log:WARNING'),
        ],
        [
            InlineKeyboardButton("🟠 HIGH+" + (" ✅" if min_severity == 'HIGH' else ""), callback_data='admin:audit_log:HIGH'),
            InlineKeyboardButton("🔴 CRITICAL" + (" ✅" if min_severity == 'CRITICAL' else ""), callback_data='admin:audit_log:CRITICAL'),
        ],
        [InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')],
    ]
    await query.edit_message_text(
        text[:4000], parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def _export_excel(query, context):
    """
    FIX جدید: خروجی کامل دیتابیس (کاربران، تیکت‌ها، آمار سوالات)
    به‌صورت یک فایل اکسل با چند شیت، آماده دانلود.
    """
    await query.edit_message_text("⏳ <b>در حال ساخت فایل اکسل...</b>", parse_mode='HTML')
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import io

        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        # ── شیت کاربران ──
        ws = wb.active
        ws.title = 'کاربران'
        headers = ['آیدی', 'نام', 'شماره دانشجویی', 'گروه', 'ورودی', 'یوزرنیم', 'وضعیت', 'تاریخ ثبت‌نام', 'تعداد پاسخ', 'پاسخ صحیح']
        ws.append(headers)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        users = await db.all_users(approved_only=False)
        for u in users:
            ws.append([
                u.get('user_id', ''), u.get('name', ''), u.get('student_id', '') or '—',
                u.get('group', ''), u.get('intake', '') or '—', u.get('username', '') or '—',
                'تأییدشده' if u.get('approved') else 'در انتظار',
                u.get('registered_at', '')[:10],
                u.get('total_answers', 0), u.get('correct_answers', 0),
            ])

        # ── شیت تیکت‌ها ──
        ws2 = wb.create_sheet('تیکت‌ها')
        headers2 = ['شماره تیکت', 'کاربر', 'موضوع', 'وضعیت', 'تاریخ ثبت']
        ws2.append(headers2)
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
        tickets = await db.tickets.find({}).sort('created_at', -1).to_list(2000)
        for t in tickets:
            ws2.append([
                t.get('ticket_id', ''), t.get('user_name', ''), t.get('subject', ''),
                'باز' if t.get('status') == 'open' else 'بسته',
                t.get('created_at', '')[:10],
            ])

        # ── شیت آمار سوالات ──
        ws3 = wb.create_sheet('بانک سوال')
        headers3 = ['درس', 'مبحث', 'سختی', 'تعداد پاسخ', 'پاسخ صحیح', 'وضعیت تأیید']
        ws3.append(headers3)
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
        questions = await db.questions.find({}).to_list(5000)
        for q in questions:
            ws3.append([
                q.get('lesson', ''), q.get('topic', ''), q.get('difficulty', ''),
                q.get('attempt_count', 0), q.get('correct_count', 0),
                'تأییدشده' if q.get('approved') else 'در انتظار',
            ])

        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
                sheet.column_dimensions[col[0].column_letter].width = min(max_len + 3, 35)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        fname = f"humsyar_export_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        await query.message.reply_document(
            document=buf, filename=fname,
            caption=f"📥 <b>خروجی کامل دیتابیس</b>\n👥 {len(users)} کاربر | 🎫 {len(tickets)} تیکت | 🧪 {len(questions)} سوال",
            parse_mode='HTML'
        )
        # FIX جدید طبق سند: خروجی گرفتن دیتابیس شامل اطلاعات حساس
        # کاربران است — باید ثبت شود که کِی و توسط کدام مدیر دانلود شد.
        admin_user_doc = await db.get_user(uid)
        actor_name = admin_user_doc.get('name', 'مدیر ارشد') if admin_user_doc else 'مدیر ارشد'
        actor_role = await db.get_actor_role_label(uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, uid,
            "خروجی اکسل دیتابیس", module='Backup', severity='HIGH',
            actor_role=actor_role,
            details=f"{len(users)} کاربر | {len(tickets)} تیکت | {len(questions)} سوال",
            tags=['خروجی_اکسل']
        )
        await query.edit_message_text(
            "✅ فایل اکسل ارسال شد!",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')]])
        )
    except Exception as e:
        logger.error(f"_export_excel error: {e}")
        await query.edit_message_text(
            f"❌ خطا در ساخت فایل: {e}",
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')]])
        )


async def _show_intakes(query):
    intakes = await db.get_all_intakes()
    keyboard = []
    for i in intakes:
        code  = i['code']
        label = i['label']
        icon  = "✅" if i.get('active', True) else "❌"
        keyboard.append([
            InlineKeyboardButton(f"{icon} {label}", callback_data=f'admin:intake_view:{code}'),
            InlineKeyboardButton("🔄", callback_data=f'admin:intake_toggle:{code}'),
            InlineKeyboardButton("🗑", callback_data=f'admin:intake_del:{code}'),
        ])
    keyboard.append([InlineKeyboardButton("➕ افزودن ورودی جدید", callback_data='admin:intake_add')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')])
    await query.edit_message_text(
        "📅 <b>مدیریت ورودی‌های دانشجویی</b>\n\n━━━━━━━━━━━━━━━━\n✅=فعال | ❌=غیرفعال | 🔄=تغییر | 🗑=حذف",
        parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


# ══════════════════════════════════════════════════
# 🛡 سطوح دسترسی چندگانه ادمین — UI
# ══════════════════════════════════════════════════

async def _show_roles(query):
    """لیست نقش‌های فرعی فعلی + دکمه افزودن"""
    roles = await db.get_all_admin_roles()
    keyboard = []
    if not roles:
        text = (
            "🛡 <b>سطوح دسترسی ادمین</b>\n\n"
            "━━━━━━━━━━━━━━━━\n"
            "هنوز هیچ نقش فرعی تعریف نشده.\n\n"
            "💡 شما (مدیر ارشد) همیشه دسترسی کامل دارید.\n"
            "می‌توانید برای دیگران نقش محدودتر بسازید:\n"
            "• 🎫 پشتیبان — فقط پاسخ به تیکت\n"
            "• 🎓 مدیر محتوا — کلی یا محدود به یک ورودی\n"
            "• 📢 مسئول اطلاعیه — فقط ارسال همگانی"
        )
    else:
        text = "🛡 <b>سطوح دسترسی ادمین</b>\n\n━━━━━━━━━━━━━━━━\n"
        for r in roles:
            target_uid = r['_id']
            u = await db.get_user(target_uid)
            name = u.get('name', '') if u else f"آیدی {target_uid}"
            role_label = db.ROLE_LABELS.get(r.get('role', ''), r.get('role', ''))
            scope = r.get('scope_intake')
            scope_txt = f" ({scope})" if scope else ""
            text += f"\n👤 <b>{name}</b>\n   {role_label}{scope_txt}\n"
            keyboard.append([InlineKeyboardButton(
                f"🗑 حذف نقش {name}", callback_data=f'admin:role_remove:{target_uid}'
            )])
    keyboard.append([InlineKeyboardButton("➕ افزودن نقش جدید", callback_data='admin:role_add_pick')])
    keyboard.append([InlineKeyboardButton("🔙 بازگشت به پنل", callback_data='admin:main')])
    await query.edit_message_text(text, parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))


async def _show_role_type_picker(query):
    """انتخاب نوع نقش قبل از گرفتن آیدی کاربر"""
    keyboard = [
        [InlineKeyboardButton("🎫 پشتیبان (فقط تیکت)", callback_data='admin:role_type:support')],
        [InlineKeyboardButton("🎓 مدیر محتوا (کلی)",   callback_data='admin:role_type:content_admin')],
        [InlineKeyboardButton("📅 مدیر محتوا (محدود به ورودی)", callback_data='admin:role_type:content_scoped')],
        [InlineKeyboardButton("📢 مسئول اطلاعیه",        callback_data='admin:role_type:broadcaster')],
        [InlineKeyboardButton("🔙 بازگشت", callback_data='admin:roles')],
    ]
    await query.edit_message_text(
        "🛡 <b>افزودن نقش جدید</b>\n\nنوع نقش را انتخاب کنید:",
        parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard)
    )


async def _show_role_intake_picker(query):
    """برای content_scoped — انتخاب ورودی که این مدیر فقط به آن دسترسی دارد"""
    intakes = await db.get_all_intakes()
    if not intakes:
        await query.answer("❌ هنوز هیچ ورودی‌ای تعریف نشده! اول از «مدیریت ورودی‌ها» اضافه کنید.", show_alert=True)
        return
    keyboard = [[InlineKeyboardButton(i['label'], callback_data=f'admin:role_intake:{i["code"]}')] for i in intakes]
    keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:role_add_pick')])
    await query.edit_message_text(
        "📅 <b>مدیر محتوای محدود</b>\n\nاین مدیر فقط به محتوای کدام ورودی دسترسی داشته باشد؟",
        parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard)
    )


# ══════════════════════════════════════════════════
# هندلرهای متن
# ══════════════════════════════════════════════════

async def handle_admin_text(update: Update, context: ContextTypes.DEFAULT_TYPE) -> bool:
    mode = context.user_data.get('mode', '')
    text = update.message.text.strip()

    # FIX جدید: متن دلخواه حالت تعمیر و نگهداری
    if mode == 'add_required_channel':
        context.user_data['mode'] = ''
        parts_txt = [p.strip() for p in text.split(',', 1)]
        if len(parts_txt) < 2 or not parts_txt[0].lstrip('-').isdigit():
            await update.message.reply_text(
                "❌ فرمت اشتباه!\nمثال: <code>-1001234567890, نام کانال</code>",
                parse_mode='HTML'
            )
            return True
        ch_id, ch_title = parts_txt[0], parts_txt[1]
        ok = await db.add_required_channel(ch_id, ch_title)
        if ok:
            admin_user = await db.get_user(uid)
            actor_name = admin_user.get('name', 'مدیر ارشد') if admin_user else 'مدیر ارشد'
            actor_role = await db.get_actor_role_label(uid)
            await send_audit_log(
                context.bot, 'admin', actor_name, uid,
                "افزودن کانال اجباری", module='Settings', severity='HIGH',
                actor_role=actor_role,
                target_id=ch_id, target_type='channel', target_label=ch_title,
                tags=['کانال_اجباری']
            )
            await update.message.reply_text(
                f"✅ کانال <b>{ch_title}</b> اضافه شد.\n\n"
                "⚠️ مطمئن شوید ربات در این کانال ادمین است.",
                parse_mode='HTML',
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔒 بازگشت به مدیریت کانال", callback_data='admin:channel_lock')
                ]])
            )
        else:
            await update.message.reply_text("⚠️ این کانال قبلاً اضافه شده.")
        return True

    if mode == 'set_maintenance_text':
        context.user_data['mode'] = ''
        if text in ('پیشفرض', 'پیش‌فرض', '-'):
            await db.set_setting('maintenance_text', '')
            msg = "✅ متن حالت تعمیر به پیش‌فرض بازگشت."
        else:
            await db.set_setting('maintenance_text', text)
            msg = "✅ متن حالت تعمیر ذخیره شد."
        await update.message.reply_text(msg, reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("⚙️ بازگشت به تنظیمات", callback_data='admin:settings')
        ]]))
        return True

    # FIX جدید: تنظیم گروه لاگ — هم آیدی عددی، هم پیام فوروارد‌شده
    if mode in ('set_log_group_admin', 'set_log_group_content'):
        key = 'log_group_admin' if mode == 'set_log_group_admin' else 'log_group_content'
        context.user_data['mode'] = ''
        if text in ('حذف', '-'):
            await db.set_setting(key, None)
            await update.message.reply_text(
                "✅ تنظیم گروه حذف شد.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⚙️ بازگشت به تنظیمات", callback_data='admin:settings')]])
            )
            return True
        chat_id = None
        # حالت ۱: پیام فوروارد‌شده از طرف خودِ گروه (پیام anonymous گروه)
        fwd = update.message.forward_origin
        if fwd is not None and hasattr(fwd, 'sender_chat') and fwd.sender_chat is not None:
            chat_id = fwd.sender_chat.id
        # حالت ۲ (قابل‌اعتمادترین روش): آیدی عددی منفی مستقیم گروه
        # — با فوروارد یک پیام به @RawDataBot یا @userinfobot پیدا می‌شود
        elif text.lstrip('-').isdigit():
            chat_id = int(text)
        if not chat_id:
            await update.message.reply_text(
                "⚠️ آیدی عددی گروه (با علامت منفی، مثلاً <code>-1001234567890</code>) را بفرستید.\n\n"
                "💡 برای پیدا کردن آیدی گروه: یک پیام از آن گروه را به @RawDataBot فوروارد کنید "
                "و مقدار <code>chat.id</code> را از پاسخ آن کپی کنید.",
                parse_mode='HTML'
            )
            return True
        await db.set_setting(key, chat_id)
        try:
            await context.bot.send_message(chat_id, "✅ این گروه به‌عنوان گروه لاگ ربات تنظیم شد.")
        except Exception:
            await update.message.reply_text(
                "⚠️ گروه ذخیره شد، اما ربات نتوانست پیام تستی بفرستد — مطمئن شوید ربات عضو آن گروه است."
            )
        await update.message.reply_text(
            f"✅ آیدی گروه <code>{chat_id}</code> ذخیره شد.", parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("⚙️ بازگشت به تنظیمات", callback_data='admin:settings')]])
        )
        return True

    # FIX جدید: گام نهایی افزودن نقش فرعی ادمین — گرفتن آیدی عددی
    if mode == 'add_admin_role':
        if not text.isdigit():
            await update.message.reply_text(
                "⚠️ آیدی باید فقط عدد باشد. دوباره وارد کنید یا /cancel بزنید."
            )
            return True
        target_uid   = int(text)
        role_type    = context.user_data.pop('new_role_type', '')
        scope_intake = context.user_data.pop('new_role_intake', None)
        context.user_data['mode'] = ''
        admin_uid = update.effective_user.id
        ok = await db.add_admin_role(target_uid, role_type, admin_uid, scope_intake)
        if not ok:
            await update.message.reply_text("❌ نوع نقش نامعتبر است.")
            return True
        role_label = db.ROLE_LABELS.get(role_type, role_type)
        scope_txt  = f" — محدود به ورودی {scope_intake}" if scope_intake else ""
        admin_user = await db.get_user(admin_uid)
        actor_name = admin_user.get('name', 'مدیر ارشد') if admin_user else 'مدیر ارشد'
        actor_role = await db.get_actor_role_label(uid)
        target_user_for_log = await db.get_user(target_uid)
        await send_audit_log(
            context.bot, 'admin', actor_name, admin_uid,
            "انتساب رول به کاربر", module='Roles', severity='CRITICAL',
            actor_role=actor_role,
            target_id=str(target_uid), target_type='user',
            target_label=target_user_for_log.get('name', '') if target_user_for_log else '',
            details=f"نقش: {role_label}{scope_txt}",
            tags=['انتساب_نقش']
        )
        try:
            await context.bot.send_message(
                target_uid,
                f"🛡 <b>دسترسی جدید به شما داده شد!</b>\n\nنقش: {role_label}{scope_txt}\n\n"
                "از دکمه «👨‍⚕️ پنل ادمین» در منوی اصلی استفاده کنید.",
                parse_mode='HTML'
            )
        except Exception:
            pass
        await update.message.reply_text(
            f"✅ نقش <b>{role_label}</b>{scope_txt} برای آیدی <code>{target_uid}</code> ثبت شد.",
            parse_mode='HTML',
            reply_markup=InlineKeyboardMarkup([[
                InlineKeyboardButton("🛡 بازگشت به سطوح دسترسی", callback_data='admin:roles')
            ]])
        )
        return True

    if mode == 'search_user':
        users = await db.search_users(text)
        context.user_data['mode'] = ''
        if not users:
            await update.message.reply_text(f"❌ کاربری با «{text}» پیدا نشد.",
                reply_markup=InlineKeyboardMarkup([[
                    InlineKeyboardButton("🔍 جستجوی مجدد", callback_data='admin:search_user'),
                    InlineKeyboardButton("🔙 پنل ادمین", callback_data='admin:main'),
                ]]))
            return True
        keyboard = [[InlineKeyboardButton(
            f"{'✅' if u.get('approved') else '⏳'} {u.get('name','')} | {u.get('student_id','') or u.get('username','N/A')}",
            callback_data=f'admin:user_detail:{u["user_id"]}')] for u in users]
        keyboard.append([InlineKeyboardButton("🔙 بازگشت", callback_data='admin:main')])
        await update.message.reply_text(f"🔍 <b>{len(users)} نتیجه برای «{text}»:</b>", parse_mode='HTML', reply_markup=InlineKeyboardMarkup(keyboard))
        return True

    elif mode == 'edit_user':
        info  = context.user_data.get('edit_user', {})
        uid   = info.get('uid')
        field = info.get('field')
        label = info.get('label', '')
        if uid and field:
            await db.update_user(uid, {field: text})
            context.user_data['mode'] = ''
            await update.message.reply_text(f"✅ {label} ویرایش شد.",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("👤 مشاهده کاربر", callback_data=f'admin:user_detail:{uid}')]]))
            return True

    elif mode == 'add_intake':
        try:
            pts = [p.strip() for p in text.split(',', 1)]
            if len(pts) < 2:
                raise ValueError("فرمت اشتباه")
            code, label = pts[0], pts[1]
            ok = await db.add_intake(code, label)
            context.user_data.pop('mode', None)
            if ok:
                await update.message.reply_text(f"✅ ورودی <b>{label}</b> اضافه شد!", parse_mode='HTML',
                    reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("📅 مدیریت ورودی‌ها", callback_data='admin:intakes')]]))
            else:
                await update.message.reply_text(f"⚠️ ورودی با کد <code>{code}</code> قبلاً وجود دارد.", parse_mode='HTML')
            return True
        except ValueError:
            await update.message.reply_text("❌ فرمت اشتباه!\nمثال: <code>bahman_1404, بهمن ۱۴۰۴</code>", parse_mode='HTML')
            return True

    elif mode == 'qbank_awaiting_desc':
        desc     = '' if text == '-' else text
        lesson   = context.user_data.get('qbank_lesson', '')
        topic    = context.user_data.get('qbank_topic', '')
        file_id  = context.user_data.get('qbank_file_id', '')
        ftype    = context.user_data.get('qbank_file_type', 'document')
        if file_id:
            await db.add_qbank_file(lesson, topic, file_id, desc, ftype)
            context.user_data['mode'] = ''
            await update.message.reply_text(f"✅ فایل بانک سوال اضافه شد!\n📚 {lesson} — {topic}",
                reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("🔙 بازگشت به بانک سوال", callback_data='admin:qbank_manage')]]))
        return True

    return False


async def upload_file_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    if uid != ADMIN_ID:
        return
    if context.user_data.get('mode') != 'qbank_awaiting_file':
        return
    doc = update.message.document or (update.message.photo[-1] if update.message.photo else None)
    if not doc:
        await update.message.reply_text("❌ فایل معتبر ارسال کنید.")
        return
    context.user_data.update({
        'qbank_file_id':   doc.file_id,
        'qbank_file_type': 'photo' if update.message.photo else 'document',
        'mode':            'qbank_awaiting_desc',
    })
    lesson = context.user_data.get('qbank_lesson', '')
    topic  = context.user_data.get('qbank_topic', '')
    await update.message.reply_text(
        f"✅ فایل دریافت شد!\n📚 {lesson} — {topic}\n\n📝 توضیح کوتاه وارد کنید (یا <code>-</code> بزنید):",
        parse_mode='HTML',
        reply_markup=InlineKeyboardMarkup([[InlineKeyboardButton("❌ لغو", callback_data='admin:qbank_manage')]]))


def get_keyboard_for_uid(user, uid: int):
    if uid == ADMIN_ID:
        return admin_keyboard()
    role = user.get('role', 'student') if user else 'student'
    if role == 'content_admin':
        return content_admin_keyboard()
    return main_keyboard()
